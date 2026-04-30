#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OsonSavdo - Advanced Marketplace Telegram Bot
Uzum + Glovo + CRM + Support + Delivery System
"""

import asyncio
import json
import os
import io
import logging
import sqlite3
import random
import string
from datetime import datetime, timedelta, time as dtime
from collections import deque
from typing import Optional

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    InputFile, ReplyKeyboardRemove
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes, ConversationHandler,
    PicklePersistence
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ─── CONFIG ────────────────────────────────────────────────────────────────────
BOT_TOKEN = "8609713083:AAFoh_EZqps4cSIs7sdTqdoWpFBox_Z-C80"
ADMIN_ID = 7948989650
DB_PATH = "/data/oson_savdo.db"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ─── CONVERSATION STATES ───────────────────────────────────────────────────────
(
    WAITING_ADDRESS, WAITING_PAYMENT_SCREENSHOT, WAITING_PROMO,
    WAITING_SHOP_NAME, WAITING_SHOP_DESC, WAITING_PRODUCT_NAME,
    WAITING_PRODUCT_DESC, WAITING_PRODUCT_PRICE, WAITING_PRODUCT_PHOTO,
    WAITING_PRODUCT_CATEGORY, WAITING_TICKET_MSG, WAITING_TICKET_REPLY,
    WAITING_REVIEW, WAITING_COURIER_NAME, WAITING_OPERATOR_ORDER,
    WAITING_PROMO_CODE_CREATE, WAITING_PRODUCT_STOCK, WAITING_DELIVERY_PRICE,
    WAITING_WORK_HOURS, WAITING_DISCOUNT_PERCENT, WAITING_COMMISSION,
    WAITING_BROADCAST, WAITING_PRODUCT_EDIT_PRICE, WAITING_REFERRAL,
    WAITING_SHOP_EDIT, WAITING_PRODUCT_DISCOUNT,
    WAITING_JOB_SHOP_NAME, WAITING_JOB_SHOP_DESC,
    WAITING_ADMIN_SHOP_NAME, WAITING_ADMIN_SHOP_DESC, WAITING_ADMIN_SHOP_OWNER,
    WAITING_CARD_NUMBER, WAITING_CARD_HOLDER,
    WAITING_SUB_PERCENT,
    WAITING_PROMO_CODE, WAITING_PROMO_TYPE, WAITING_PROMO_VALUE,
    WAITING_PROMO_LIMIT, WAITING_PROMO_DAYS, WAITING_PROMO_MIN_AMOUNT,
    WAITING_PHONE_NAME, WAITING_PHONE_NUMBER,
    WAITING_TEL_ORDER_ITEMS, WAITING_TEL_ORDER_ADDRESS,
) = range(44)

# ─── DATABASE ──────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    c.executescript("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    );

    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY,
        tg_id INTEGER UNIQUE,
        username TEXT,
        full_name TEXT,
        role TEXT DEFAULT 'customer',
        referral_code TEXT UNIQUE,
        referred_by INTEGER,
        total_orders INTEGER DEFAULT 0,
        total_spent REAL DEFAULT 0,
        rating REAL DEFAULT 0,
        bonus_balance REAL DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS shops (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner_tg_id INTEGER,
        name TEXT,
        description TEXT,
        category TEXT DEFAULT 'Umumiy',
        status TEXT DEFAULT 'pending',
        delivery_price REAL DEFAULT 0,
        work_hours TEXT DEFAULT '09:00-22:00',
        rating REAL DEFAULT 0,
        total_reviews INTEGER DEFAULT 0,
        card_number TEXT DEFAULT '',
        card_holder TEXT DEFAULT '',
        is_open INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shop_id INTEGER,
        name TEXT,
        description TEXT,
        price REAL,
        discount_percent REAL DEFAULT 0,
        photo_id TEXT,
        category TEXT DEFAULT 'Umumiy',
        stock INTEGER DEFAULT 999,
        is_active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_tg_id INTEGER,
        shop_id INTEGER,
        items TEXT,
        address TEXT,
        payment_method TEXT,
        payment_status TEXT DEFAULT 'pending',
        payment_screenshot TEXT,
        promo_code TEXT,
        discount_amount REAL DEFAULT 0,
        subtotal REAL DEFAULT 0,
        delivery_price REAL DEFAULT 0,
        total REAL DEFAULT 0,
        commission REAL DEFAULT 0,
        courier_tg_id INTEGER,
        courier_type TEXT DEFAULT 'standard',
        premium_fee REAL DEFAULT 0,
        status TEXT DEFAULT 'new',
        operator_note TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS couriers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tg_id INTEGER UNIQUE,
        name TEXT,
        is_premium INTEGER DEFAULT 0,
        is_busy INTEGER DEFAULT 0,
        is_active INTEGER DEFAULT 1,
        total_deliveries INTEGER DEFAULT 0,
        rating REAL DEFAULT 0,
        last_assigned TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS courier_queue (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER,
        added_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS promo_codes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT UNIQUE,
        discount_type TEXT,
        discount_value REAL,
        max_uses INTEGER DEFAULT 100,
        used_count INTEGER DEFAULT 0,
        expires_at TEXT,
        is_active INTEGER DEFAULT 1,
        min_order_amount REAL DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS favorites (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_tg_id INTEGER,
        shop_id INTEGER,
        UNIQUE(user_tg_id, shop_id)
    );

    CREATE TABLE IF NOT EXISTS reviews (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_tg_id INTEGER,
        shop_id INTEGER,
        order_id INTEGER,
        rating INTEGER,
        comment TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS tickets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_tg_id INTEGER,
        subject TEXT,
        status TEXT DEFAULT 'open',
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS ticket_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ticket_id INTEGER,
        sender_tg_id INTEGER,
        message TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS operator_orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        operator_tg_id INTEGER,
        user_tg_id INTEGER,
        raw_request TEXT,
        converted_order TEXT,
        status TEXT DEFAULT 'pending',
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_tg_id INTEGER,
        message TEXT,
        is_read INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS shop_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shop_id INTEGER UNIQUE,
        fee_percent REAL DEFAULT 0,
        last_paid_at TEXT,
        next_due_at TEXT,
        total_earned REAL DEFAULT 0,
        status TEXT DEFAULT 'active'
    );
    """)

    # Default settings
    defaults = [
        ("commission_percent", "10"),
        ("premium_courier_fee", "15000"),
        ("referral_bonus", "5000"),
        ("platform_name", "OsonSavdo"),
        ("order_phones", ""),  # JSON list: [{"name":"Ism","phone":"+998901234567","shop_id":0}]
    ]
    for key, val in defaults:
        c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (key, val))

    conn.commit()
    conn.close()

def migrate_db():
    """Eski DB ga yangi ustunlar qo'shish"""
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("ALTER TABLE shops ADD COLUMN card_number TEXT DEFAULT ''")
    except:
        pass
    try:
        c.execute("ALTER TABLE shops ADD COLUMN card_holder TEXT DEFAULT ''")
    except:
        pass
    try:
        c.execute("ALTER TABLE promo_codes ADD COLUMN min_order_amount REAL DEFAULT 0")
    except:
        pass
    try:
        c.execute("ALTER TABLE shops ADD COLUMN is_open INTEGER DEFAULT 1")
    except:
        pass
    # NULL qiymatlarni tuzatish
    c.execute("UPDATE shops SET is_open=1 WHERE is_open IS NULL")
    try:
        c.execute("""CREATE TABLE IF NOT EXISTS shop_subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_id INTEGER UNIQUE,
            fee_percent REAL DEFAULT 0,
            last_paid_at TEXT,
            next_due_at TEXT,
            total_earned REAL DEFAULT 0,
            status TEXT DEFAULT 'active'
        )""")
    except:
        pass
    conn.commit()
    conn.close()

def get_setting(key: str, default: str = "") -> str:
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row["value"] if row else default

def set_setting(key: str, value: str):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()

# ─── HELPERS ───────────────────────────────────────────────────────────────────
def gen_referral_code(tg_id: int) -> str:
    return f"REF{tg_id}"

def get_or_create_user(tg_id: int, username: str = "", full_name: str = "") -> dict:
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    if not user:
        ref_code = gen_referral_code(tg_id)
        conn.execute(
            "INSERT INTO users (tg_id, username, full_name, referral_code) VALUES (?, ?, ?, ?)",
            (tg_id, username, full_name, ref_code)
        )
        conn.commit()
        user = conn.execute("SELECT * FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()
    return dict(user)

def get_user_role(tg_id: int) -> str:
    if tg_id == ADMIN_ID:
        return "admin"
    conn = get_db()
    row = conn.execute("SELECT role FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()
    if row:
        return row["role"]
    return "customer"

def format_price(amount: float) -> str:
    return f"{amount:,.0f} so'm"

def order_status_emoji(status: str) -> str:
    emojis = {
        "new": "🆕", "confirmed": "✅", "rejected": "❌",
        "courier_assigned": "🚴", "delivering": "📦",
        "delivered": "🏁", "cancelled": "🚫"
    }
    return emojis.get(status, "❓")

def order_status_text(status: str) -> str:
    texts = {
        "new": "Yangi", "confirmed": "Tasdiqlangan", "rejected": "Rad etilgan",
        "courier_assigned": "Kuryer tayinlandi", "delivering": "Yetkazilmoqda",
        "delivered": "Yetkazildi", "cancelled": "Bekor qilindi"
    }
    return texts.get(status, status)

# ─── INLINE KEYBOARDS ──────────────────────────────────────────────────────────
def main_menu_kb(role: str) -> InlineKeyboardMarkup:
    buttons = []
    if role in ("customer", "admin", "operator"):
        buttons += [
            [InlineKeyboardButton("🏪 Do'konlar", callback_data="shops_list"),
             InlineKeyboardButton("🛒 Savat", callback_data="cart_view")],
            [InlineKeyboardButton("📦 Buyurtmalarim", callback_data="my_orders"),
             InlineKeyboardButton("👤 Profil", callback_data="profile")],
            [InlineKeyboardButton("❤️ Sevimlilar", callback_data="favorites"),
             InlineKeyboardButton("🎫 Promo kod", callback_data="promo_enter")],
            [InlineKeyboardButton("🎟 Ticket ochish", callback_data="ticket_open"),
             InlineKeyboardButton("🔗 Referal", callback_data="referral")],
            [InlineKeyboardButton("💼 Ishga kirish", callback_data="job_apply")],
        ]
    if role == "shop_owner":
        buttons += [
            [InlineKeyboardButton("🏪 Do'konim", callback_data="my_shop"),
             InlineKeyboardButton("📦 Buyurtmalar", callback_data="shop_orders")],
            [InlineKeyboardButton("📊 Hisobot", callback_data="shop_report"),
             InlineKeyboardButton("➕ Mahsulot", callback_data="add_product")],
        ]
    if role == "courier":
        buttons += [
            [InlineKeyboardButton("🚴 Mening buyurtmalarim", callback_data="courier_orders"),
             InlineKeyboardButton("📊 Statistika", callback_data="courier_stats")],
        ]
    if role == "admin":
        buttons += [
            [InlineKeyboardButton("⚙️ Admin Panel", callback_data="admin_panel")],
        ]
    if role == "operator":
        buttons += [
            [InlineKeyboardButton("🎙 Operator Panel", callback_data="operator_panel")],
        ]
    return InlineKeyboardMarkup(buttons)

def back_kb(callback: str = "main_menu") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Orqaga", callback_data=callback)]])

# ─── /START ────────────────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    username = update.effective_user.username or ""
    full_name = update.effective_user.full_name or ""

    # Referral handling
    args = context.args
    if args and args[0].startswith("REF"):
        ref_code = args[0]
        conn = get_db()
        referrer = conn.execute("SELECT * FROM users WHERE referral_code=?", (ref_code,)).fetchone()
        conn.close()
        if referrer and referrer["tg_id"] != tg_id:
            context.user_data["referred_by"] = referrer["tg_id"]

    user = get_or_create_user(tg_id, username, full_name)

    if "referred_by" in context.user_data:
        ref_by = context.user_data.pop("referred_by")
        conn = get_db()
        existing = conn.execute("SELECT referred_by FROM users WHERE tg_id=?", (tg_id,)).fetchone()
        if existing and not existing["referred_by"]:
            bonus = float(get_setting("referral_bonus", "5000"))
            conn.execute("UPDATE users SET referred_by=? WHERE tg_id=?", (ref_by, tg_id))
            conn.execute("UPDATE users SET bonus_balance=bonus_balance+? WHERE tg_id=?", (bonus, ref_by))
            conn.commit()
            try:
                await context.bot.send_message(ref_by, f"🎉 Referal bonus! +{format_price(bonus)} hisob balansiga qo'shildi!")
            except:
                pass
        conn.close()

    role = get_user_role(tg_id)
    platform = get_setting("platform_name", "OsonSavdo")

    # Do'konlarni yuklash
    conn = get_db()
    shops = conn.execute("SELECT * FROM shops WHERE status='approved' ORDER BY rating DESC").fetchall()
    user_row = conn.execute("SELECT bonus_balance FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()

    bonus = user_row["bonus_balance"] if user_row else 0
    cart = get_cart(context)
    cart_count = sum(v["qty"] for v in cart.values())
    cart_label = f"🛒 Savat ({cart_count})" if cart_count else "🛒 Savat"

    if role == "customer" or role == "admin" or role == "operator":
        if not shops:
            text = (
                f"👋 <b>{full_name}</b>, xush kelibsiz!\n"
                f"🛍 <b>{platform}</b>\n\n"
                f"🏪 Hozircha do'konlar yo'q."
            )
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(cart_label, callback_data="cart_view")],
                [InlineKeyboardButton("☰ Ko'proq", callback_data="more_menu")],
            ])
        else:
            text = f"🛍 <b>{platform}</b> — Do'konlar\n\n"
            buttons = []
            for s in shops:
                is_open = s["is_open"] if "is_open" in s.keys() else 1
                status_icon = "🟢" if is_open else "🔴"
                rating = f"⭐{s['rating']:.1f}" if s["rating"] else "⭐"
                buttons.append([InlineKeyboardButton(
                    f"{status_icon} {s['name']} {rating} | 🚚{format_price(s['delivery_price'])}",
                    callback_data=f"shop_{s['id']}"
                )])
            buttons.append([
                InlineKeyboardButton(cart_label, callback_data="cart_view"),
                InlineKeyboardButton("☰ Ko'proq", callback_data="more_menu"),
            ])
            kb = InlineKeyboardMarkup(buttons)
    else:
        role_names = {"shop_owner": "Do'kon egasi", "courier": "Kuryer", "admin": "Admin", "operator": "Operator"}
        text = (
            f"👋 Xush kelibsiz, <b>{full_name}</b>!\n"
            f"🛍 <b>{platform}</b>\n\n"
            f"👤 Rolingiz: <b>{role_names.get(role, role)}</b>"
        )
        kb = main_menu_kb(role)

    if update.message:
        await update.message.reply_text(text, reply_markup=kb, parse_mode="HTML")
    else:
        await update.callback_query.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

# ─── SHOPS LIST ────────────────────────────────────────────────────────────────
async def more_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id
    role = get_user_role(tg_id)

    conn = get_db()
    user_row = conn.execute("SELECT bonus_balance FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()
    bonus = user_row["bonus_balance"] if user_row else 0

    buttons = [
        [InlineKeyboardButton("📞 Telefon orqali buyurtma", callback_data="phone_order")],
        [InlineKeyboardButton("📦 Buyurtmalarim", callback_data="my_orders"),
         InlineKeyboardButton("👤 Profil", callback_data="profile")],
        [InlineKeyboardButton("❤️ Sevimlilar", callback_data="favorites"),
         InlineKeyboardButton("🔗 Referal", callback_data="referral")],
        [InlineKeyboardButton("🎟 Murojaat", callback_data="ticket_open"),
         InlineKeyboardButton("💼 Ishga kirish", callback_data="job_apply")],
    ]
    if bonus > 0:
        buttons.append([InlineKeyboardButton(f"💎 Bonus: {format_price(bonus)}", callback_data="profile")])
    if role == "admin":
        buttons.append([InlineKeyboardButton("⚙️ Admin Panel", callback_data="admin_panel")])
    if role == "operator":
        buttons.append([InlineKeyboardButton("🎙 Operator Panel", callback_data="operator_panel")])
    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")])

    await q.edit_message_text(
        "☰ <b>Qo'shimcha imkoniyatlar</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

# ─── TELEFON ORQALI BUYURTMA (MIJOZ) ──────────────────────────────────────────
async def phone_order_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []

    if not phones:
        await q.edit_message_text(
            "📞 <b>Telefon orqali buyurtma</b>\n\n"
            "⚠️ Hozircha telefon raqam belgilanmagan.\n"
            "Iltimos, bot orqali buyurtma bering.",
            parse_mode="HTML",
            reply_markup=back_kb("more_menu")
        )
        return

    text = (
        "📞 <b>Telefon orqali buyurtma</b>\n\n"
        "Quyidagi raqamlardan biriga qo'ng'iroq qiling:\n\n"
    )
    for p in phones:
        shop_name = f" ({p.get('shop_name', '')})" if p.get('shop_name') else ""
        text += f"👤 <b>{p['name']}</b>{shop_name}\n📱 <code>{p['phone']}</code>\n\n"

    text += "📋 Do'kon egasi sizning buyurtmangizni qabul qilib, kuryerga uzatadi."

    await q.edit_message_text(
        text,
        parse_mode="HTML",
        reply_markup=back_kb("more_menu")
    )

async def shops_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    shops = conn.execute(
        "SELECT * FROM shops WHERE status='approved' ORDER BY rating DESC"
    ).fetchall()
    conn.close()

    if not shops:
        await q.edit_message_text(
            "🏪 Hozircha do'konlar yo'q.\nKo'proq vaqt o'tgach qaytib keling!",
            reply_markup=back_kb()
        )
        return

    buttons = []
    for s in shops:
        stars = "⭐" * int(s["rating"]) if s["rating"] else "⭐"
        is_open = s["is_open"] if "is_open" in s.keys() else 1
        closed_mark = " 🔴" if not is_open else ""
        buttons.append([InlineKeyboardButton(
            f"🏪 {s['name']}{closed_mark} {stars} | 🚚 {format_price(s['delivery_price'])}",
            callback_data=f"shop_{s['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")])
    await q.edit_message_text(
        "🏪 <b>Do'konlar ro'yxati</b>\n\nBiror do'konni tanlang:",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def shop_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[1])

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()

    if not shop:
        await q.edit_message_text("Do'kon topilmadi.", reply_markup=back_kb("shops_list"))
        return

    context.user_data["current_shop"] = shop_id
    tg_id = update.effective_user.id

    # Check if favorite
    conn = get_db()
    fav = conn.execute(
        "SELECT id FROM favorites WHERE user_tg_id=? AND shop_id=?", (tg_id, shop_id)
    ).fetchone()
    conn.close()

    fav_btn = "💔 Sevimlilardan chiqar" if fav else "❤️ Sevimlilarga qo'sh"
    fav_cb = f"unfav_{shop_id}" if fav else f"fav_{shop_id}"

    rating_str = f"⭐ {shop['rating']:.1f}" if shop['rating'] else "⭐ Reyting yo'q"
    text = (
        f"🏪 <b>{shop['name']}</b>\n"
        f"📝 {shop['description']}\n"
        f"📂 Kategoriya: {shop['category']}\n"
        f"⏰ Ish vaqti: {shop['work_hours']}\n"
        f"🚚 Yetkazish: {format_price(shop['delivery_price'])}\n"
        f"{rating_str} ({shop['total_reviews']} sharh)\n"
    )

    # Bu do'konn telefon raqamlari
    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []
    shop_phones = [p for p in phones if p.get("shop_id") == shop_id]

    phone_buttons = []
    if shop_phones:
        phone_text = "\n📞 <b>Telefon orqali buyurtma:</b>\n"
        for p in shop_phones:
            phone_text += f"👤 {p['name']}: <code>{p['phone']}</code>\n"
        phone_buttons = [[InlineKeyboardButton("📞 Telefon orqali buyurtma", callback_data=f"shop_call_{shop_id}")]]
    else:
        phone_text = ""

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📦 Mahsulotlar", callback_data=f"products_{shop_id}")],
        [InlineKeyboardButton(fav_btn, callback_data=fav_cb)],
        [InlineKeyboardButton("⭐ Reyting qo'y", callback_data=f"rate_shop_{shop_id}"),
         InlineKeyboardButton("🛒 Savat", callback_data="cart_view")],
        *phone_buttons,
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="shops_list")],
    ])
    await q.edit_message_text(text + phone_text, reply_markup=kb, parse_mode="HTML")

# ─── MIJOZ TELEFON BUYURTMA ───────────────────────────────────────────────────
async def shop_call_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])

    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []
    shop_phones = [p for p in phones if p.get("shop_id") == shop_id]

    conn = get_db()
    shop = conn.execute("SELECT name FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()
    shop_name = shop["name"] if shop else ""

    if not shop_phones:
        await q.edit_message_text(
            "📞 <b>Telefon orqali buyurtma</b>\n\n"
            "⚠️ Bu do\'kon uchun telefon raqam belgilanmagan.\n"
            "Iltimos, bot orqali buyurtma bering.",
            parse_mode="HTML",
            reply_markup=back_kb(f"shop_{shop_id}")
        )
        return

    text = f"📞 <b>{shop_name} — Telefon buyurtma</b>\n\nQuyidagi raqamga qo\'ng\'iroq qiling:\n\n"
    for p in shop_phones:
        text += f"👤 <b>{p['name']}</b>\n📱 <code>{p['phone']}</code>\n\n"
    text += "📋 Operator sizning buyurtmangizni qabul qilib, kuryerga uzatadi."

    await q.edit_message_text(
        text,
        parse_mode="HTML",
        reply_markup=back_kb(f"shop_{shop_id}")
    )

# ─── PRODUCTS ──────────────────────────────────────────────────────────────────
async def products_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[1])

    conn = get_db()
    products = conn.execute(
        "SELECT * FROM products WHERE shop_id=? AND is_active=1", (shop_id,)
    ).fetchall()
    shop = conn.execute("SELECT name, is_open FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()

    if not products:
        await q.edit_message_text(
            "📦 Bu do'konda mahsulotlar yo'q.",
            reply_markup=back_kb(f"shop_{shop_id}")
        )
        return

    is_open = shop["is_open"] if shop and "is_open" in shop.keys() else 1

    buttons = []
    for p in products:
        price = p["price"] * (1 - p["discount_percent"] / 100) if p["discount_percent"] else p["price"]
        disc = f" 🔥-{int(p['discount_percent'])}%" if p["discount_percent"] else ""
        stock_txt = f" ({p['stock']} ta)" if p["stock"] < 10 else ""
        buttons.append([InlineKeyboardButton(
            f"🛒 {p['name']}{disc} — {format_price(price)}{stock_txt}",
            callback_data=f"add_cart_{p['id']}"
        )])

    cart = get_cart(context)
    cart_count = sum(v["qty"] for v in cart.values())
    cart_label = f"🛒 Savat ({cart_count})" if cart_count else "🛒 Savat"

    buttons.append([
        InlineKeyboardButton(cart_label, callback_data="cart_view"),
        InlineKeyboardButton("⬅️ Orqaga", callback_data=f"shop_{shop_id}"),
    ])

    status = "🟢 Ochiq" if is_open else "🔴 Yopiq"
    await q.edit_message_text(
        f"📦 <b>{shop['name'] if shop else ''}</b> — mahsulotlar\n"
        f"{status}\n\n"
        f"<i>Mahsulotga bosing — savatga tushadi!</i>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def product_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    product_id = int(q.data.split("_")[1])

    conn = get_db()
    p = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    conn.close()

    if not p:
        await q.answer("Mahsulot topilmadi", show_alert=True)
        return

    price = p["price"]
    disc_text = ""
    if p["discount_percent"]:
        discounted = price * (1 - p["discount_percent"] / 100)
        disc_text = f"\n🔥 Aksiya: <s>{format_price(price)}</s> → <b>{format_price(discounted)}</b> (-{int(p['discount_percent'])}%)"
        price = discounted

    text = (
        f"🛍 <b>{p['name']}</b>\n"
        f"📝 {p['description']}\n"
        f"💰 Narx: <b>{format_price(price)}</b>{disc_text}\n"
        f"📦 Omborda: {p['stock']} ta\n"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🛒 Savatga qo'sh", callback_data=f"add_cart_{product_id}")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data=f"products_{p['shop_id']}")],
    ])

    if p["photo_id"]:
        try:
            await context.bot.send_photo(
                q.from_user.id, p["photo_id"], caption=text,
                reply_markup=kb, parse_mode="HTML"
            )
            await q.message.delete()
        except:
            await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")
    else:
        await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

# ─── CART ──────────────────────────────────────────────────────────────────────
def get_cart(context) -> dict:
    return context.user_data.get("cart", {})

def save_cart(context, cart: dict):
    context.user_data["cart"] = cart

async def add_to_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    product_id = int(q.data.split("_")[2])

    conn = get_db()
    p = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    if not p:
        await q.answer("❌ Mahsulot topilmadi!", show_alert=True)
        conn.close()
        return

    shop = conn.execute("SELECT is_open FROM shops WHERE id=?", (p["shop_id"],)).fetchone()
    conn.close()

    is_open = shop["is_open"] if shop and "is_open" in shop.keys() else 1
    if not is_open:
        await q.answer("🔴 Do'kon hozir yopiq! Keyinroq urinib ko'ring.", show_alert=True)
        return

    await q.answer(f"✅ {p['name']} savatga qo'shildi!")

    cart = get_cart(context)
    pid = str(product_id)
    if pid in cart:
        cart[pid]["qty"] += 1
    else:
        price = p["price"] * (1 - p["discount_percent"] / 100) if p["discount_percent"] else p["price"]
        cart[pid] = {
            "name": p["name"], "price": price,
            "qty": 1, "shop_id": p["shop_id"]
        }
    save_cart(context, cart)

    # Mahsulotlar sahifasini yangilash (savat soni ko'rinsin)
    cart_count = sum(v["qty"] for v in cart.values())
    cart_label = f"🛒 Savat ({cart_count})"
    shop_id = p["shop_id"]

    conn2 = get_db()
    products = conn2.execute("SELECT * FROM products WHERE shop_id=? AND is_active=1", (shop_id,)).fetchall()
    shop2 = conn2.execute("SELECT name, is_open FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn2.close()

    buttons = []
    for pr in products:
        price2 = pr["price"] * (1 - pr["discount_percent"] / 100) if pr["discount_percent"] else pr["price"]
        disc2 = f" 🔥-{int(pr['discount_percent'])}%" if pr["discount_percent"] else ""
        stock_txt = f" ({pr['stock']} ta)" if pr["stock"] < 10 else ""
        buttons.append([InlineKeyboardButton(
            f"🛒 {pr['name']}{disc2} — {format_price(price2)}{stock_txt}",
            callback_data=f"add_cart_{pr['id']}"
        )])

    is_open2 = shop2["is_open"] if shop2 and "is_open" in shop2.keys() else 1
    status2 = "🟢 Ochiq" if is_open2 else "🔴 Yopiq"
    buttons.append([
        InlineKeyboardButton(cart_label, callback_data="cart_view"),
        InlineKeyboardButton("⬅️ Orqaga", callback_data=f"shop_{shop_id}"),
    ])

    try:
        await q.edit_message_text(
            f"📦 <b>{shop2['name'] if shop2 else ''}</b> — mahsulotlar\n"
            f"{status2}\n\n"
            f"<i>Mahsulotga bosing — savatga tushadi!</i>",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode="HTML"
        )
    except:
        pass

async def cart_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    cart = get_cart(context)

    if not cart:
        await q.edit_message_text(
            "🛒 Savat bo'sh!\n\nDo'konlardan mahsulot qo'shing.",
            reply_markup=back_kb("shops_list")
        )
        return

    total = sum(v["price"] * v["qty"] for v in cart.values())
    text = "🛒 <b>Savatingiz:</b>\n\n"
    buttons = []

    for pid, item in cart.items():
        text += f"• {item['name']} x{item['qty']} = {format_price(item['price'] * item['qty'])}\n"
        buttons.append([
            InlineKeyboardButton(f"➖", callback_data=f"cart_dec_{pid}"),
            InlineKeyboardButton(f"{item['name'][:15]} x{item['qty']}", callback_data="noop"),
            InlineKeyboardButton(f"➕", callback_data=f"cart_inc_{pid}"),
            InlineKeyboardButton(f"🗑", callback_data=f"cart_del_{pid}"),
        ])

    # Check promo
    promo_discount = context.user_data.get("promo_discount", 0)
    promo_code = context.user_data.get("promo_code", "")
    if promo_discount:
        text += f"\n🏷 Promo: -{format_price(promo_discount)}"
        total -= promo_discount

    text += f"\n\n💰 <b>Jami: {format_price(total)}</b>"

    buttons += [
        [InlineKeyboardButton("🎫 Promo kod", callback_data="promo_enter"),
         InlineKeyboardButton("🗑 Tozalash", callback_data="cart_clear")],
        [InlineKeyboardButton("🛍 Do'konlarga qaytish", callback_data="main_menu")],
        [InlineKeyboardButton("✅ ZAKAZ BERISH", callback_data="checkout")],
    ]

    await q.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode="HTML")

async def cart_update(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    parts = q.data.split("_")
    action = parts[1]
    pid = parts[2]

    cart = get_cart(context)
    if pid not in cart:
        await q.answer("⚠️ Savat yangilandi, qayta oching", show_alert=True)
        await cart_view(update, context)
        return

    if action == "inc":
        cart[pid]["qty"] += 1
    elif action == "dec":
        cart[pid]["qty"] -= 1
        if cart[pid]["qty"] <= 0:
            del cart[pid]
    elif action == "del":
        del cart[pid]

    save_cart(context, cart)
    await q.answer("✅")

    # Savat bo'm-bo'sh bo'lib qolsa
    if not cart:
        await q.edit_message_text(
            "🛒 Savat bo'sh!\n\nDo'konlardan mahsulot qo'shing.",
            reply_markup=back_kb("shops_list")
        )
        return

    await cart_view(update, context)

async def cart_clear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["cart"] = {}
    context.user_data.pop("promo_discount", None)
    context.user_data.pop("promo_code", None)
    await q.edit_message_text("🗑 Savat tozalandi.", reply_markup=back_kb("shops_list"))

# ─── PROMO CODE ────────────────────────────────────────────────────────────────
async def promo_enter(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["state"] = WAITING_PROMO
    await q.edit_message_text(
        "🎫 Promo kodni kiriting:",
        reply_markup=back_kb("cart_view")
    )
    return WAITING_PROMO

async def promo_apply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    code = update.message.text.strip().upper()
    conn = get_db()
    promo = conn.execute(
        "SELECT * FROM promo_codes WHERE code=? AND is_active=1", (code,)
    ).fetchone()
    conn.close()

    if not promo:
        await update.message.reply_text("❌ Promo kod topilmadi yoki faol emas.")
        return ConversationHandler.END

    now = datetime.now()
    if promo["expires_at"] and datetime.fromisoformat(promo["expires_at"]) < now:
        await update.message.reply_text("❌ Promo kod muddati tugagan.")
        return ConversationHandler.END

    if promo["used_count"] >= promo["max_uses"]:
        await update.message.reply_text("❌ Promo kod limitiga yetildi.")
        return ConversationHandler.END

    cart = get_cart(context)
    subtotal = sum(v["price"] * v["qty"] for v in cart.values())

    min_amount = promo["min_order_amount"] if promo["min_order_amount"] else 0
    if min_amount > 0 and subtotal < min_amount:
        await update.message.reply_text(
            f"❌ Bu promo kod faqat <b>{format_price(min_amount)}</b> dan yuqori buyurtmalarda ishlaydi.\n"
            f"Sizning savatcha: <b>{format_price(subtotal)}</b>",
            parse_mode="HTML"
        )
        return ConversationHandler.END

    if promo["discount_type"] == "percent":
        discount = subtotal * promo["discount_value"] / 100
    else:
        discount = promo["discount_value"]

    context.user_data["promo_discount"] = discount
    context.user_data["promo_code"] = code

    await update.message.reply_text(
        f"✅ Promo kod qo'llanildi!\n💰 Chegirma: -{format_price(discount)}",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🛒 Savatga", callback_data="cart_view")]])
    )
    return ConversationHandler.END

# ─── CHECKOUT ──────────────────────────────────────────────────────────────────
async def checkout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    cart = get_cart(context)

    if not cart:
        await q.edit_message_text("🛒 Savat bo'sh!", reply_markup=back_kb("shops_list"))
        return

    context.user_data["checkout_step"] = "address"
    await q.edit_message_text(
        "📍 <b>Yetkazish manzilini kiriting:</b>\n\n"
        "Masalan: Toshkent sh., Chilonzor t., 5-uy, 12-xonadon",
        reply_markup=back_kb("cart_view"),
        parse_mode="HTML"
    )
    return WAITING_ADDRESS

async def got_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    address = update.message.text.strip()
    context.user_data["order_address"] = address
    tg_id = update.effective_user.id

    # Bonus balansini tekshir
    conn = get_db()
    user = conn.execute("SELECT bonus_balance FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()
    bonus_bal = user["bonus_balance"] if user else 0

    buttons = [
        [InlineKeyboardButton("💳 Karta orqali", callback_data="pay_card"),
         InlineKeyboardButton("💵 Naqd", callback_data="pay_cash")],
    ]
    if bonus_bal >= 1000:
        buttons.append([
            InlineKeyboardButton(f"💎 Bonus hisobdan ({format_price(bonus_bal)})", callback_data="pay_bonus")
        ])

    await update.message.reply_text(
        f"📍 Manzil: <b>{address}</b>\n\n"
        f"💳 To'lov usulini tanlang:"
        + (f"\n\n💎 Bonus hisobingiz: <b>{format_price(bonus_bal)}</b>" if bonus_bal >= 1000 else ""),
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )
    return ConversationHandler.END

async def payment_method(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if q.data == "pay_bonus":
        context.user_data["payment_method"] = "bonus"
        await place_order(update, context, screenshot=None)
        return
    method = "card" if q.data == "pay_card" else "cash"
    context.user_data["payment_method"] = method

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🚴 Oddiy kuryer (bepul)", callback_data="courier_standard"),
         InlineKeyboardButton("⚡ Premium kuryer", callback_data="courier_premium")],
    ])

    premium_fee = get_setting("premium_courier_fee", "15000")
    await q.edit_message_text(
        f"✅ To'lov: <b>{'Karta' if method == 'card' else 'Naqd'}</b>\n\n"
        f"🚴 Kuryer turini tanlang:\n"
        f"• Oddiy: Bepul\n"
        f"• Premium: +{format_price(float(premium_fee))} (tezroq yetkazish)",
        reply_markup=kb,
        parse_mode="HTML"
    )

async def courier_type_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ctype = "standard" if q.data == "courier_standard" else "premium"
    context.user_data["courier_type"] = ctype

    if context.user_data.get("payment_method") == "card":
        cart = get_cart(context)
        total = sum(v["price"] * v["qty"] for v in cart.values())
        promo = context.user_data.get("promo_discount", 0)
        total -= promo

        # Get shop delivery price and card info
        shop_id = list(cart.values())[0]["shop_id"] if cart else None
        card_info_text = ""
        if shop_id:
            conn = get_db()
            shop = conn.execute("SELECT delivery_price, card_number, card_holder FROM shops WHERE id=?", (shop_id,)).fetchone()
            conn.close()
            delivery = shop["delivery_price"] if shop else 0
            if shop and shop["card_number"]:
                cn = shop["card_number"]
                formatted_card = " ".join([cn[i:i+4] for i in range(0, len(cn), 4)])
                card_info_text = (
                    f"\n\n💳 <b>To'lov kartasi:</b>\n"
                    f"🔢 Raqam: <code>{formatted_card}</code>\n"
                    f"👤 Ism: <b>{shop['card_holder'] or ''}</b>\n\n"
                    f"⬆️ Ushbu kartaga o'tkazma qiling, so'ng chekni yuboring."
                )
            else:
                card_info_text = "\n\n⚠️ Do'kon karta ma'lumotlarini hali kiritмagan. Aloqa qiling."
        else:
            delivery = 0

        premium_fee = float(get_setting("premium_courier_fee", "15000")) if ctype == "premium" else 0
        grand_total = total + delivery + premium_fee

        await q.edit_message_text(
            f"💳 <b>Karta orqali to'lov</b>\n\n"
            f"💰 Jami: <b>{format_price(grand_total)}</b>"
            f"{card_info_text}\n\n"
            f"📸 To'lov chekini (screenshot) yuboring:",
            parse_mode="HTML"
        )
        return WAITING_PAYMENT_SCREENSHOT
    else:
        await place_order(update, context, screenshot=None)

async def got_payment_screenshot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.photo:
        file_id = update.message.photo[-1].file_id
    elif update.message.document:
        file_id = update.message.document.file_id
    else:
        await update.message.reply_text("❌ Iltimos, rasm yoki fayl yuboring.")
        return WAITING_PAYMENT_SCREENSHOT

    await place_order(update, context, screenshot=file_id)
    return ConversationHandler.END

async def place_order(update, context, screenshot=None):
    tg_id = update.effective_user.id
    cart = get_cart(context)

    if not cart:
        return

    # Get shop info
    shop_id = list(cart.values())[0]["shop_id"]
    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()

    subtotal = sum(v["price"] * v["qty"] for v in cart.values())
    promo_discount = context.user_data.get("promo_discount", 0)
    promo_code = context.user_data.get("promo_code", "")
    delivery_price = shop["delivery_price"] if shop else 0
    ctype = context.user_data.get("courier_type", "standard")
    premium_fee = float(get_setting("premium_courier_fee", "15000")) if ctype == "premium" else 0

    # Bonus balansni olish va ishlatish
    conn_b = get_db()
    user_row = conn_b.execute("SELECT bonus_balance FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    conn_b.close()
    bonus_balance = user_row["bonus_balance"] if user_row else 0
    before_bonus = subtotal - promo_discount + delivery_price + premium_fee
    bonus_used = min(bonus_balance, max(before_bonus, 0))
    total = max(before_bonus - bonus_used, 0)

    commission_pct = float(get_setting("commission_percent", "10"))
    commission = total * commission_pct / 100

    items_json = json.dumps(dict(cart), ensure_ascii=False)
    address = context.user_data.get("order_address", "")
    payment_method = context.user_data.get("payment_method", "cash")
    payment_status = "pending" if payment_method == "card" else "cash"

    conn = get_db()
    cur = conn.execute(
        """INSERT INTO orders (user_tg_id, shop_id, items, address, payment_method,
           payment_status, payment_screenshot, promo_code, discount_amount,
           subtotal, delivery_price, total, commission, courier_type, premium_fee)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (tg_id, shop_id, items_json, address, payment_method,
         payment_status, screenshot, promo_code, promo_discount + bonus_used,
         subtotal, delivery_price, total, commission, ctype, premium_fee)
    )
    order_id = cur.lastrowid

    # Update promo usage
    if promo_code:
        conn.execute("UPDATE promo_codes SET used_count=used_count+1 WHERE code=?", (promo_code,))

    # Bonus ayir
    if bonus_used > 0:
        conn.execute("UPDATE users SET bonus_balance=bonus_balance-? WHERE tg_id=?", (bonus_used, tg_id))

    conn.commit()
    conn.close()

    # Clear cart
    context.user_data["cart"] = {}
    context.user_data.pop("promo_discount", None)
    context.user_data.pop("promo_code", None)

    # Chegirma satrlari
    discount_lines = ""
    if promo_discount > 0:
        discount_lines += f"🎫 Promo chegirma: -{format_price(promo_discount)}\n"
    if bonus_used > 0:
        discount_lines += f"💎 Bonus chegirma: -{format_price(bonus_used)}\n"

    order_text = (
        f"🆕 <b>Yangi buyurtma #{1000 + order_id}</b>\n\n"
        f"👤 Mijoz: {update.effective_user.full_name}\n"
        f"🏪 Do'kon: {shop['name'] if shop else 'N/A'}\n"
        f"📍 Manzil: {address}\n"
        f"💰 Mahsulotlar: {format_price(subtotal)}\n"
        + discount_lines +
        f"💵 Jami: <b>{format_price(total)}</b>\n"
        f"💳 To'lov: {'Karta' if payment_method == 'card' else 'Naqd'}\n"
        f"🚴 Kuryer: {'Premium ⚡' if ctype == 'premium' else 'Oddiy'}\n"
    )

    # Admin notify o'chirilgan — faqat do'kon egasiga xabar boradi

    # Notify shop owner — qabul/rad tugmalari bilan
    if shop:
        try:
            items_list = "\n".join([f"  • {v['name']} x{v['qty']} = {format_price(v['price'] * v['qty'])}" for v in cart.values()])
            shop_notify_text = (
                f"🛎 <b>Yangi buyurtma #{1000 + order_id}!</b>\n\n"
                f"👤 Mijoz: {update.effective_user.full_name}\n"
                f"📦 Mahsulotlar:\n{items_list}\n\n"
                f"📍 Manzil: {address}\n"
                f"💰 Mahsulotlar: {format_price(subtotal)}\n"
                f"🚚 Yetkazish: {format_price(delivery_price)}\n"
                f"💵 Jami: <b>{format_price(total)}</b>\n"
                f"💳 To'lov: {'Karta' if payment_method == 'card' else 'Naqd'}\n"
                f"🚴 Kuryer: {'Premium ⚡' if ctype == 'premium' else 'Oddiy'}\n\n"
                f"⚡ <b>Iltimos, tezda qabul qiling!</b>"
            )
            shop_kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Qabul qilish", callback_data=f"shop_confirm_{order_id}"),
                 InlineKeyboardButton("❌ Rad etish", callback_data=f"shop_reject_{order_id}")],
            ])
            await context.bot.send_message(
                shop["owner_tg_id"], shop_notify_text, reply_markup=shop_kb, parse_mode="HTML"
            )
            if screenshot:
                await context.bot.send_photo(
                    shop["owner_tg_id"], screenshot,
                    caption=f"📸 #{1000 + order_id} — to'lov cheki"
                )
        except:
            pass

    confirm_text = (
        f"✅ <b>Buyurtmangiz qabul qilindi!</b>\n\n"
        f"📦 Buyurtma ID: <b>#{1000 + order_id}</b>\n"
        f"💰 Jami: <b>{format_price(total)}</b>\n"
        f"⏳ Do'kon egasi tasdiqlashini kuting...\n\n"
        f"📲 Buyurtma holatini <b>Buyurtmalarim</b> bo'limida kuzating."
    )

    if update.message:
        await update.message.reply_text(
            confirm_text,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📦 Buyurtmalarim", callback_data="my_orders")]]),
            parse_mode="HTML"
        )
    else:
        await context.bot.send_message(
            tg_id, confirm_text,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📦 Buyurtmalarim", callback_data="my_orders")]]),
            parse_mode="HTML"
        )

# ─── MY ORDERS ─────────────────────────────────────────────────────────────────
async def my_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    orders = conn.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id "
        "WHERE o.user_tg_id=? ORDER BY o.id DESC LIMIT 10",
        (tg_id,)
    ).fetchall()
    conn.close()

    if not orders:
        await q.edit_message_text("📦 Buyurtmalar yo'q.", reply_markup=back_kb())
        return

    buttons = []
    for o in orders:
        emoji = order_status_emoji(o["status"])
        buttons.append([InlineKeyboardButton(
            f"{emoji} #{1000 + o['id']} — {o['shop_name']} — {format_price(o['total'])}",
            callback_data=f"order_detail_{o['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")])
    await q.edit_message_text(
        "📦 <b>Buyurtmalarim (so'nggi 10)</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def order_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])

    conn = get_db()
    o = conn.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id WHERE o.id=?",
        (order_id,)
    ).fetchone()
    conn.close()

    if not o:
        await q.answer("Buyurtma topilmadi", show_alert=True)
        return

    items = json.loads(o["items"])
    items_text = "\n".join([f"• {v['name']} x{v['qty']}" for v in items.values()])
    emoji = order_status_emoji(o["status"])
    status_txt = order_status_text(o["status"])

    text = (
        f"{emoji} <b>Buyurtma #{1000 + o['id']}</b>\n\n"
        f"🏪 Do'kon: {o['shop_name']}\n"
        f"📦 Mahsulotlar:\n{items_text}\n\n"
        f"📍 Manzil: {o['address']}\n"
        f"💰 Jami: {format_price(o['total'])}\n"
        f"💳 To'lov: {'Karta' if o['payment_method'] == 'card' else 'Naqd'}\n"
        f"📊 Holat: <b>{status_txt}</b>\n"
    )

    buttons = []
    if o["status"] == "delivered":
        buttons.append([InlineKeyboardButton("⭐ Baho berish", callback_data=f"rate_order_{order_id}")])
        buttons.append([InlineKeyboardButton("🔄 Qayta buyurtma", callback_data=f"reorder_{order_id}")])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="my_orders")])
    await q.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode="HTML")

# ─── REORDER ───────────────────────────────────────────────────────────────────
async def reorder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[1])

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.close()

    if not o:
        await q.answer("Buyurtma topilmadi", show_alert=True)
        return

    items = json.loads(o["items"])
    context.user_data["cart"] = items
    await q.edit_message_text(
        "✅ Mahsulotlar savatga qo'shildi!",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🛒 Savatga", callback_data="cart_view")]])
    )

# ─── ADMIN ORDER CONFIRM/REJECT ────────────────────────────────────────────────
async def admin_confirm_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])

    conn = get_db()
    conn.execute(
        "UPDATE orders SET status='confirmed', updated_at=datetime('now') WHERE id=?",
        (order_id,)
    )
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.commit()
    conn.close()

    await q.edit_message_text(f"✅ Buyurtma #{1000 + order_id} tasdiqlandi!")

    # Notify customer
    try:
        await context.bot.send_message(
            o["user_tg_id"],
            f"✅ Buyurtmangiz #{1000 + order_id} tasdiqlandi!\n🚴 Kuryer tayinlanmoqda..."
        )
    except:
        pass

    # Assign courier
    await assign_courier(context, order_id, o["courier_type"])

async def admin_reject_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.execute(
        "UPDATE orders SET status='rejected', updated_at=datetime('now') WHERE id=?",
        (order_id,)
    )
    conn.commit()
    conn.close()

    await q.edit_message_text(f"❌ Buyurtma #{1000 + order_id} rad etildi.")

    try:
        await context.bot.send_message(
            o["user_tg_id"],
            f"❌ Buyurtmangiz #{1000 + order_id} rad etildi.\n📞 Qo'shimcha ma'lumot uchun support bilan bog'laning."
        )
    except:
        pass

async def shop_confirm_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer("✅ Tasdiqlandi!")
    order_id = int(q.data.split("_")[2])

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if o and o["status"] == "new":
        conn.execute("UPDATE orders SET status='confirmed' WHERE id=?", (order_id,))
        conn.commit()
    conn.close()

    await q.edit_message_text(f"✅ Buyurtma #{1000 + order_id} qabul qilindi!")

    try:
        await context.bot.send_message(
            o["user_tg_id"],
            f"✅ Do'kon buyurtmangizni #{1000 + order_id} qabul qildi!\n🚴 Kuryer tayinlanmoqda..."
        )
    except:
        pass

    await assign_courier(context, order_id, o["courier_type"] if o else "standard")

async def shop_reject_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer("❌ Rad etildi!")
    order_id = int(q.data.split("_")[2])

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.execute("UPDATE orders SET status='rejected' WHERE id=?", (order_id,))
    conn.commit()
    conn.close()

    await q.edit_message_text(f"❌ Buyurtma #{1000 + order_id} rad etildi.")

    if o:
        try:
            await context.bot.send_message(
                o["user_tg_id"],
                f"❌ Buyurtmangiz #{1000 + order_id} rad etildi."
            )
        except:
            pass

# ─── COURIER SYSTEM (ROUND-ROBIN) ─────────────────────────────────────────────
courier_queue_state = deque()

async def assign_courier(context, order_id: int, courier_type: str = "standard"):
    conn = get_db()
    is_premium = 1 if courier_type == "premium" else 0

    couriers = conn.execute(
        "SELECT * FROM couriers WHERE is_active=1 AND is_busy=0 AND is_premium=? ORDER BY last_assigned ASC NULLS FIRST",
        (is_premium,)
    ).fetchall()

    if not couriers and is_premium:
        couriers = conn.execute(
            "SELECT * FROM couriers WHERE is_active=1 AND is_busy=0 ORDER BY last_assigned ASC NULLS FIRST"
        ).fetchall()

    if not couriers:
        # Add to queue
        conn.execute("INSERT INTO courier_queue (order_id) VALUES (?)", (order_id,))
        conn.commit()
        conn.close()
        try:
            o = conn.execute("SELECT user_tg_id FROM orders WHERE id=?", (order_id,)).fetchone()
            if o:
                await context.bot.send_message(
                    o["user_tg_id"],
                    f"⏳ Buyurtma #{1000 + order_id} kuryer navbatiga qo'shildi. Tez orada tayinlanadi."
                )
        except:
            pass
        return

    courier = couriers[0]
    conn.execute(
        "UPDATE orders SET courier_tg_id=?, status='courier_assigned', updated_at=datetime('now') WHERE id=?",
        (courier["tg_id"], order_id)
    )
    conn.execute(
        "UPDATE couriers SET is_busy=1, last_assigned=datetime('now') WHERE tg_id=?",
        (courier["tg_id"],)
    )
    conn.commit()
    conn.close()

    o_conn = get_db()
    o = o_conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    o_conn.close()

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Qabul qilish", callback_data=f"courier_accept_{order_id}"),
         InlineKeyboardButton("❌ Rad etish", callback_data=f"courier_reject_{order_id}")],
    ])

    try:
        items = json.loads(o["items"]) if o else {}
        items_txt = ", ".join([f"{v['name']} x{v['qty']}" for v in items.values()])
        await context.bot.send_message(
            courier["tg_id"],
            f"🚴 <b>Yangi buyurtma #{1000 + order_id}</b>\n\n"
            f"📦 {items_txt}\n"
            f"📍 Manzil: {o['address']}\n"
            f"💰 {format_price(o['total'])}\n"
            f"{'⚡ Premium yetkazish' if courier_type == 'premium' else ''}",
            reply_markup=kb,
            parse_mode="HTML"
        )
    except:
        pass

async def courier_accept(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])
    tg_id = update.effective_user.id

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.execute("UPDATE orders SET status='delivering' WHERE id=?", (order_id,))
    conn.commit()
    conn.close()

    await q.edit_message_text(f"✅ Buyurtma #{1000 + order_id} qabul qilindi!\n🚴 Yetkazishni boshlang.")

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🏁 Yetkazildi", callback_data=f"courier_delivered_{order_id}")],
    ])
    await context.bot.send_message(tg_id, "Yetkazib bo'lgach tugmani bosing:", reply_markup=kb)

    if o:
        try:
            await context.bot.send_message(
                o["user_tg_id"],
                f"🚴 Kuryer yetib bormoqda!\n📦 Buyurtma #{1000 + order_id}"
            )
        except:
            pass

async def courier_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])
    tg_id = update.effective_user.id

    conn = get_db()
    conn.execute("UPDATE couriers SET is_busy=0 WHERE tg_id=?", (tg_id,))
    conn.execute("UPDATE orders SET courier_tg_id=NULL, status='confirmed' WHERE id=?", (order_id,))
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.commit()
    conn.close()

    await q.edit_message_text(f"❌ Buyurtma #{1000 + order_id} rad etildi.")

    # Reassign
    ctype = o["courier_type"] if o else "standard"
    await assign_courier(context, order_id, ctype)

async def courier_delivered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])
    tg_id = update.effective_user.id

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.execute(
        "UPDATE orders SET status='delivered', updated_at=datetime('now') WHERE id=?",
        (order_id,)
    )
    conn.execute("UPDATE couriers SET is_busy=0, total_deliveries=total_deliveries+1 WHERE tg_id=?", (tg_id,))
    conn.commit()
    conn.close()

    await q.edit_message_text(f"🏁 Buyurtma #{1000 + order_id} yetkazildi! Rahmat!")

    if o:
        try:
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("⭐ Baho berish", callback_data=f"rate_order_{order_id}")],
            ])
            await context.bot.send_message(
                o["user_tg_id"],
                f"🎉 Buyurtmangiz #{1000 + order_id} yetkazildi!\n"
                f"Xizmatimiz sifatini baholang:",
                reply_markup=kb
            )
        except:
            pass

    # Check queue
    conn = get_db()
    queued = conn.execute("SELECT * FROM courier_queue ORDER BY id ASC LIMIT 1").fetchone()
    if queued:
        conn.execute("DELETE FROM courier_queue WHERE id=?", (queued["id"],))
        conn.commit()
        conn.close()
        await assign_courier(context, queued["order_id"])
    else:
        conn.close()

# ─── PROFILE ───────────────────────────────────────────────────────────────────
async def profile_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()

    if not user:
        await q.edit_message_text("Profil topilmadi.", reply_markup=back_kb())
        return

    ref_link = f"https://t.me/OsonSavdoBot?start={user['referral_code']}"
    text = (
        f"👤 <b>Profilim</b>\n\n"
        f"👤 Ism: {user['full_name']}\n"
        f"📦 Buyurtmalar: {user['total_orders']}\n"
        f"💰 Jami sarf: {format_price(user['total_spent'])}\n"
        f"⭐ Reyting: {user['rating']:.1f}\n"
        f"💎 Bonus balans: {format_price(user['bonus_balance'])}\n"
        f"🔗 Referal: <code>{ref_link}</code>\n"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")],
    ])
    await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

# ─── FAVORITES ─────────────────────────────────────────────────────────────────
async def favorites_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    favs = conn.execute(
        "SELECT s.* FROM favorites f JOIN shops s ON f.shop_id=s.id WHERE f.user_tg_id=?",
        (tg_id,)
    ).fetchall()
    conn.close()

    if not favs:
        await q.edit_message_text("❤️ Sevimli do'konlar yo'q.", reply_markup=back_kb())
        return

    buttons = [[InlineKeyboardButton(f"🏪 {s['name']}", callback_data=f"shop_{s['id']}")] for s in favs]
    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")])
    await q.edit_message_text("❤️ <b>Sevimli do'konlar</b>", reply_markup=InlineKeyboardMarkup(buttons), parse_mode="HTML")

async def add_favorite(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer("❤️ Sevimlilarga qo'shildi!")
    shop_id = int(q.data.split("_")[1])
    tg_id = update.effective_user.id

    conn = get_db()
    try:
        conn.execute("INSERT OR IGNORE INTO favorites (user_tg_id, shop_id) VALUES (?, ?)", (tg_id, shop_id))
        conn.commit()
    except:
        pass
    conn.close()
    await shop_detail(update, context)

async def remove_favorite(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer("💔 Sevimlilardan chiqarildi!")
    shop_id = int(q.data.split("_")[1])
    tg_id = update.effective_user.id

    conn = get_db()
    conn.execute("DELETE FROM favorites WHERE user_tg_id=? AND shop_id=?", (tg_id, shop_id))
    conn.commit()
    conn.close()
    await shop_detail(update, context)

# ─── RATING ────────────────────────────────────────────────────────────────────
async def rate_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])
    context.user_data["rating_order_id"] = order_id

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"⭐ {i}", callback_data=f"give_rating_{order_id}_{i}")
        for i in range(1, 6)
    ]])
    await q.edit_message_text("⭐ Xizmatni baholang (1-5):", reply_markup=kb)

async def give_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    parts = q.data.split("_")
    order_id = int(parts[2])
    rating = int(parts[3])

    context.user_data["pending_rating"] = {"order_id": order_id, "rating": rating}
    await q.edit_message_text(
        f"{'⭐' * rating} Baho: {rating}/5\n\n📝 Izoh yozing (yoki /skip):"
    )
    return WAITING_REVIEW

async def got_review(update: Update, context: ContextTypes.DEFAULT_TYPE):
    comment = update.message.text.strip()
    if comment == "/skip":
        comment = ""

    pending = context.user_data.pop("pending_rating", None)
    if not pending:
        return ConversationHandler.END

    order_id = pending["order_id"]
    rating = pending["rating"]
    tg_id = update.effective_user.id

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if o:
        # Check duplicate
        existing = conn.execute(
            "SELECT id FROM reviews WHERE user_tg_id=? AND order_id=?", (tg_id, order_id)
        ).fetchone()

        if not existing:
            conn.execute(
                "INSERT INTO reviews (user_tg_id, shop_id, order_id, rating, comment) VALUES (?, ?, ?, ?, ?)",
                (tg_id, o["shop_id"], order_id, rating, comment)
            )
            # Update shop rating
            avg = conn.execute(
                "SELECT AVG(rating) as avg, COUNT(*) as cnt FROM reviews WHERE shop_id=?",
                (o["shop_id"],)
            ).fetchone()
            conn.execute(
                "UPDATE shops SET rating=?, total_reviews=? WHERE id=?",
                (avg["avg"], avg["cnt"], o["shop_id"])
            )
            conn.commit()

    conn.close()
    await update.message.reply_text(
        f"⭐ Rahmat! {'⭐' * rating} bahongiz saqlandi.",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Bosh sahifa", callback_data="main_menu")]])
    )
    return ConversationHandler.END

# ─── RATE SHOP (from shop detail) ─────────────────────────────────────────────
async def rate_shop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])
    context.user_data["rating_shop_id"] = shop_id

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"⭐ {i}", callback_data=f"shop_rating_{shop_id}_{i}")
        for i in range(1, 6)
    ]])
    await q.edit_message_text("⭐ Do'konni baholang (1-5):", reply_markup=kb)

async def give_shop_rating(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer("⭐ Bahoingiz saqlandi!")
    parts = q.data.split("_")
    shop_id = int(parts[2])
    rating = int(parts[3])
    tg_id = update.effective_user.id

    conn = get_db()
    conn.execute(
        "INSERT INTO reviews (user_tg_id, shop_id, order_id, rating, comment) VALUES (?, ?, NULL, ?, '')",
        (tg_id, shop_id, rating)
    )
    avg = conn.execute(
        "SELECT AVG(rating) as avg, COUNT(*) as cnt FROM reviews WHERE shop_id=?", (shop_id,)
    ).fetchone()
    conn.execute("UPDATE shops SET rating=?, total_reviews=? WHERE id=?", (avg["avg"], avg["cnt"], shop_id))
    conn.commit()
    conn.close()

    await q.edit_message_text(
        f"{'⭐' * rating} Rahmat! Bahoyingiz saqlandi.",
        reply_markup=back_kb(f"shop_{shop_id}")
    )

# ─── TICKETS ───────────────────────────────────────────────────────────────────
async def ticket_open(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    await q.edit_message_text(
        "🎟 <b>Ticket (muammo) ochish</b>\n\n"
        "Muammoni qisqacha yozing:",
        reply_markup=back_kb(),
        parse_mode="HTML"
    )
    return WAITING_TICKET_MSG

async def got_ticket_msg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message.text.strip()
    tg_id = update.effective_user.id

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO tickets (user_tg_id, subject) VALUES (?, ?)", (tg_id, msg[:100])
    )
    ticket_id = cur.lastrowid
    conn.execute(
        "INSERT INTO ticket_messages (ticket_id, sender_tg_id, message) VALUES (?, ?, ?)",
        (ticket_id, tg_id, msg)
    )
    conn.commit()
    conn.close()

    # Notify admin
    try:
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("📩 Javob berish", callback_data=f"ticket_reply_{ticket_id}")
        ]])
        await context.bot.send_message(
            ADMIN_ID,
            f"🎟 <b>Yangi Ticket #SUP-{1000 + ticket_id}</b>\n"
            f"👤 Foydalanuvchi: {update.effective_user.full_name}\n"
            f"📝 {msg}",
            reply_markup=kb,
            parse_mode="HTML"
        )
    except:
        pass

    await update.message.reply_text(
        f"✅ Ticketingiz qabul qilindi!\n🎟 Ticket ID: <b>#SUP-{1000 + ticket_id}</b>\n\nAdmin tez orada javob beradi.",
        parse_mode="HTML",
        reply_markup=back_kb()
    )
    return ConversationHandler.END

async def ticket_reply_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    ticket_id = int(q.data.split("_")[2])
    context.user_data["replying_ticket"] = ticket_id

    await q.edit_message_text(
        f"📩 #SUP-{1000 + ticket_id} uchun javob yozing:"
    )
    return WAITING_TICKET_REPLY

async def got_ticket_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message.text.strip()
    tg_id = update.effective_user.id
    ticket_id = context.user_data.pop("replying_ticket", None)

    if not ticket_id:
        return ConversationHandler.END

    conn = get_db()
    ticket = conn.execute("SELECT * FROM tickets WHERE id=?", (ticket_id,)).fetchone()
    conn.execute(
        "INSERT INTO ticket_messages (ticket_id, sender_tg_id, message) VALUES (?, ?, ?)",
        (ticket_id, tg_id, msg)
    )
    conn.commit()
    conn.close()

    if ticket:
        try:
            await context.bot.send_message(
                ticket["user_tg_id"],
                f"📩 <b>#SUP-{1000 + ticket_id}</b> Ticket javob:\n\n{msg}",
                parse_mode="HTML"
            )
        except:
            pass

    await update.message.reply_text("✅ Javob yuborildi!")
    return ConversationHandler.END

# ─── REFERRAL ──────────────────────────────────────────────────────────────────
async def referral_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE tg_id=?", (tg_id,)).fetchone()
    referred = conn.execute("SELECT COUNT(*) as cnt FROM users WHERE referred_by=?", (tg_id,)).fetchone()
    conn.close()

    bonus = get_setting("referral_bonus", "5000")
    ref_link = f"https://t.me/OsonSavdoBot?start={user['referral_code']}"

    text = (
        f"🔗 <b>Referal tizimi</b>\n\n"
        f"Har bir do'st uchun: +{format_price(float(bonus))}\n"
        f"Taklif qilinganlar: {referred['cnt']} kishi\n\n"
        f"Sizning havolangiz:\n<code>{ref_link}</code>\n\n"
        f"Havolani do'stlarga yuboring va bonus yig'ing! 💰"
    )

    await q.edit_message_text(text, reply_markup=back_kb(), parse_mode="HTML")

# ─── SHOP OWNER PANEL ──────────────────────────────────────────────────────────
async def my_shop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE owner_tg_id=?", (tg_id,)).fetchone()
    conn.close()

    if not shop:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ Do'kon qo'shish", callback_data="add_shop")],
            [InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")],
        ])
        await q.edit_message_text(
            "🏪 Sizda do'kon yo'q.\nYangi do'kon qo'shing:",
            reply_markup=kb
        )
        return

    status_txt = {"approved": "✅ Faol", "pending": "⏳ Kutilmoqda", "rejected": "❌ Rad etildi"}.get(shop["status"], shop["status"])
    is_open = shop["is_open"] if "is_open" in shop.keys() else 1
    open_txt = "🟢 Ochiq" if is_open else "🔴 Yopiq"
    toggle_label = "🔴 Do'konni yopish" if is_open else "🟢 Do'konni ochish"

    text = (
        f"🏪 <b>{shop['name']}</b>\n\n"
        f"📝 {shop['description']}\n"
        f"📊 Holat: {status_txt}\n"
        f"🚦 Holat: {open_txt}\n"
        f"⭐ Reyting: {shop['rating']:.1f} ({shop['total_reviews']} sharh)\n"
        f"🚚 Yetkazish: {format_price(shop['delivery_price'])}\n"
        f"⏰ Ish vaqti: {shop['work_hours']}\n"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(toggle_label, callback_data=f"toggle_shop_{shop['id']}")],
        [InlineKeyboardButton("📦 Mahsulotlar", callback_data=f"owner_products_{shop['id']}"),
         InlineKeyboardButton("➕ Mahsulot qo'sh", callback_data="add_product")],
        [InlineKeyboardButton("⚙️ Sozlamalar", callback_data=f"shop_settings_{shop['id']}"),
         InlineKeyboardButton("📊 Hisobot", callback_data="shop_report")],
        [InlineKeyboardButton("📞 Telefon buyurtma kiritish", callback_data=f"tel_order_new_{shop['id']}")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")],
    ])
    await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

async def tel_order_new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[3])
    context.user_data["tel_order_shop_id"] = shop_id

    await q.edit_message_text(
        "📞 <b>Telefon orqali buyurtma kiritish</b>\n\n"
        "<b>1-qadam / 2</b>\n\n"
        "📋 Mijoz aytgan mahsulotlar va miqdorlarini yozing:\n\n"
        "<i>Masalan:\n"
        "Pepsi 1L x2\n"
        "Sut 1L x1\n"
        "Non x3</i>",
        parse_mode="HTML",
        reply_markup=back_kb("my_shop")
    )
    return WAITING_TEL_ORDER_ITEMS

async def got_tel_order_items(update: Update, context: ContextTypes.DEFAULT_TYPE):
    items_text = update.message.text.strip()
    context.user_data["tel_order_items"] = items_text

    await update.message.reply_text(
        f"✅ Mahsulotlar:\n<pre>{items_text}</pre>\n\n"
        f"<b>2-qadam / 2</b>\n\n"
        f"📍 Yetkazish manzilini yozing:",
        parse_mode="HTML"
    )
    return WAITING_TEL_ORDER_ADDRESS

async def got_tel_order_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    address = update.message.text.strip()
    shop_id = context.user_data.get("tel_order_shop_id")
    items_text = context.user_data.get("tel_order_items", "")

    # Tasdiqlash sahifasini ko'rsatish (hali saqlamay)
    context.user_data["tel_order_address"] = address

    summary = (
        f"📞 <b>Telefon buyurtma — tasdiqlang</b>\n\n"
        f"📋 <b>Mahsulotlar:</b>\n<pre>{items_text}</pre>\n\n"
        f"📍 <b>Manzil:</b> {address}\n"
        f"💵 <b>To'lov:</b> Naqd\n\n"
        f"✅ Tasdiqlaysizmi?"
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"tel_confirm_{shop_id}"),
         InlineKeyboardButton("❌ Bekor", callback_data="my_shop")],
    ])
    await update.message.reply_text(summary, reply_markup=kb, parse_mode="HTML")
    return ConversationHandler.END

async def tel_order_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])
    owner_tg_id = update.effective_user.id

    address = context.user_data.pop("tel_order_address", "")
    items_text = context.user_data.pop("tel_order_items", "")
    context.user_data.pop("tel_order_shop_id", None)

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()

    cur = conn.execute(
        """INSERT INTO orders (user_tg_id, shop_id, items, address, payment_method,
           payment_status, subtotal, delivery_price, total, commission, courier_type, premium_fee, status)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (owner_tg_id, shop_id,
         json.dumps({"tel": {"name": items_text, "price": 0, "qty": 1, "shop_id": shop_id}}),
         address, "cash", "cash", 0, 0, 0, 0, "standard", 0, "confirmed")
    )
    order_id = cur.lastrowid
    conn.commit()
    conn.close()

    order_text = (
        f"📞 <b>Telefon buyurtma #{1000 + order_id}</b>\n\n"
        f"🏪 Do'kon: {shop['name'] if shop else ''}\n"
        f"📋 Mahsulotlar:\n<pre>{items_text}</pre>\n"
        f"📍 Manzil: {address}\n"
        f"💵 To'lov: Naqd\n"
    )

    await q.edit_message_text(
        f"✅ <b>Telefon buyurtma #{1000 + order_id} saqlandi!</b>\n\n"
        f"{order_text}\n"
        f"🚴 Kuryer tayinlanmoqda...",
        parse_mode="HTML",
        reply_markup=back_kb("my_shop")
    )

    # Kuryer tayinlash (do'kon egasi o'zi tasdiqladi, admin kerak emas)
    await assign_courier(context, order_id, "standard")

async def toggle_shop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])
    tg_id = update.effective_user.id

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=? AND owner_tg_id=?", (shop_id, tg_id)).fetchone()
    if not shop:
        await q.answer("❌ Ruxsat yo'q!", show_alert=True)
        conn.close()
        return

    new_state = 0 if (shop["is_open"] if "is_open" in shop.keys() else 1) else 1
    conn.execute("UPDATE shops SET is_open=? WHERE id=?", (new_state, shop_id))
    conn.commit()
    conn.close()

    state_txt = "🟢 Ochiq" if new_state else "🔴 Yopiq"
    await q.answer(f"Do'kon {state_txt} qilindi!", show_alert=True)

    # my_shop sahifasini yangilash
    context.user_data["_refresh"] = True
    await my_shop(update, context)

async def shop_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()

    card_info = ""
    if shop and shop["card_number"]:
        card_info = f"\n\n💳 Joriy karta: <b>{shop['card_number']}</b>\n👤 Ism: <b>{shop['card_holder'] or 'Kiritilmagan'}</b>"

    # Bu do'konga tegishli telefon raqamlar
    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []
    shop_phones = [(i, p) for i, p in enumerate(phones) if p.get("shop_id") == shop_id]

    phone_info = ""
    phone_del_buttons = []
    if shop_phones:
        phone_info = "\n\n📞 <b>Telefon raqamlar:</b>\n"
        for local_idx, (_, p) in enumerate(shop_phones):
            phone_info += f"👤 {p['name']}: <code>{p['phone']}</code>\n"
            phone_del_buttons.append([InlineKeyboardButton(
                f"🗑 O'chirish: {p['name']} ({p['phone']})",
                callback_data=f"shop_del_phone_{shop_id}_{local_idx}"
            )])

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("💰 Yetkazish narxi", callback_data=f"set_delivery_{shop_id}"),
         InlineKeyboardButton("⏰ Ish vaqti", callback_data=f"set_hours_{shop_id}")],
        [InlineKeyboardButton("💳 To'lov tizimini tahrirlash", callback_data=f"set_payment_{shop_id}")],
        [InlineKeyboardButton("📞 Telefon raqam qo'shish", callback_data=f"shop_phone_{shop_id}")],
        *phone_del_buttons,
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="my_shop")],
    ])
    await q.edit_message_text(
        f"⚙️ Do'kon sozlamalari:{card_info}{phone_info}",
        reply_markup=kb,
        parse_mode="HTML"
    )

async def set_delivery_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])
    context.user_data["setting_shop_id"] = shop_id
    await q.edit_message_text("💰 Yangi yetkazish narxini kiriting (so'mda):")
    return WAITING_DELIVERY_PRICE

async def got_delivery_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        price = float(update.message.text.strip().replace(" ", ""))
    except:
        await update.message.reply_text("❌ Noto'g'ri format. Raqam kiriting.")
        return WAITING_DELIVERY_PRICE

    shop_id = context.user_data.pop("setting_shop_id", None)
    if shop_id:
        conn = get_db()
        conn.execute("UPDATE shops SET delivery_price=? WHERE id=?", (price, shop_id))
        conn.commit()
        conn.close()

    await update.message.reply_text(
        f"✅ Yetkazish narxi {format_price(price)} ga o'zgartirildi!",
        reply_markup=back_kb("my_shop")
    )
    return ConversationHandler.END

async def set_work_hours(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])
    context.user_data["setting_shop_id"] = shop_id

    conn = get_db()
    shop = conn.execute("SELECT work_hours FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()
    current = shop["work_hours"] if shop else "09:00-22:00"

    await q.edit_message_text(
        f"⏰ <b>Ish vaqtini sozlash</b>\n\n"
        f"📌 Joriy vaqt: <b>{current}</b>\n\n"
        f"Yangi ish vaqtini kiriting:\n"
        f"<i>Format: <code>HH:MM-HH:MM</code>\n"
        f"Masalan: <code>07:00-23:00</code>\n"
        f"Tungi (ertasiga): <code>22:00-02:00</code></i>",
        parse_mode="HTML",
        reply_markup=back_kb(f"shop_settings_{shop_id}")
    )
    return WAITING_WORK_HOURS

async def got_work_hours(update: Update, context: ContextTypes.DEFAULT_TYPE):
    hours = update.message.text.strip()
    shop_id = context.user_data.pop("setting_shop_id", None)

    # Format tekshiruvi
    import re
    if not re.match(r"^\d{1,2}:\d{2}-\d{1,2}:\d{2}$", hours):
        await update.message.reply_text(
            "❌ Noto'g'ri format!\n"
            "To'g'ri format: <code>07:00-23:00</code>",
            parse_mode="HTML"
        )
        return WAITING_WORK_HOURS

    try:
        open_t, close_t = hours.split("-")
        oh, om = map(int, open_t.split(":"))
        ch, cm = map(int, close_t.split(":"))
        if not (0 <= oh <= 23 and 0 <= om <= 59 and 0 <= ch <= 23 and 0 <= cm <= 59):
            raise ValueError
    except:
        await update.message.reply_text(
            "❌ Noto'g'ri vaqt! Soat 0-23, daqiqa 0-59 bo'lishi kerak.",
            parse_mode="HTML"
        )
        return WAITING_WORK_HOURS

    if shop_id:
        conn = get_db()
        conn.execute("UPDATE shops SET work_hours=? WHERE id=?", (hours, shop_id))
        conn.commit()
        conn.close()

    await update.message.reply_text(
        f"✅ <b>Ish vaqti saqlandi!</b>\n\n"
        f"⏰ <b>{hours}</b>\n\n"
        f"Bot avtomatik ravishda:\n"
        f"• <b>{open_t}</b> da do'konni ochadi 🟢\n"
        f"• <b>{close_t}</b> da do'konni yopadi 🔴",
        parse_mode="HTML",
        reply_markup=back_kb("my_shop")
    )
    return ConversationHandler.END

# ─── TO'LOV TIZIMI SOZLAMALARI ─────────────────────────────────────────────────
async def set_payment_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    parts = q.data.split("_")
    shop_id = int(parts[2])
    context.user_data["payment_shop_id"] = shop_id

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()

    current = ""
    if shop and shop["card_number"]:
        current = (
            f"\n\n📌 Joriy ma'lumot:\n"
            f"💳 Karta: <b>{shop['card_number']}</b>\n"
            f"👤 Ism: <b>{shop['card_holder'] or 'Kiritilmagan'}</b>"
        )

    await q.edit_message_text(
        f"💳 <b>To'lov tizimini sozlash</b>{current}\n\n"
        f"Karta raqamini kiriting (16 raqam, bo'shliqlarsiz):\n"
        f"Masalan: <code>8600123456789012</code>",
        parse_mode="HTML",
        reply_markup=back_kb(f"shop_settings_{shop_id}")
    )
    return WAITING_CARD_NUMBER

async def got_card_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    number = update.message.text.strip().replace(" ", "").replace("-", "")
    if not number.isdigit() or len(number) < 13 or len(number) > 19:
        await update.message.reply_text(
            "❌ Noto'g'ri karta raqami. 13-19 ta raqam kiriting.\n"
            "Masalan: <code>8600123456789012</code>",
            parse_mode="HTML"
        )
        return WAITING_CARD_NUMBER

    context.user_data["new_card_number"] = number
    await update.message.reply_text(
        f"✅ Karta raqami: <b>{number}</b>\n\n"
        f"👤 Endi karta egasining ismini kiriting (to'liq ism):\n"
        f"Masalan: <code>ALISHER KARIMOV</code>",
        parse_mode="HTML"
    )
    return WAITING_CARD_HOLDER

async def got_card_holder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    holder = update.message.text.strip().upper()
    if len(holder) < 3:
        await update.message.reply_text("❌ Ism juda qisqa. To'liq ismni kiriting.")
        return WAITING_CARD_HOLDER

    shop_id = context.user_data.pop("payment_shop_id", None)
    card_number = context.user_data.pop("new_card_number", "")

    if shop_id:
        conn = get_db()
        conn.execute(
            "UPDATE shops SET card_number=?, card_holder=? WHERE id=?",
            (card_number, holder, shop_id)
        )
        conn.commit()
        conn.close()

    # Format card number with spaces for display
    formatted = " ".join([card_number[i:i+4] for i in range(0, len(card_number), 4)])

    await update.message.reply_text(
        f"✅ <b>To'lov tizimi yangilandi!</b>\n\n"
        f"💳 Karta: <b>{formatted}</b>\n"
        f"👤 Ism: <b>{holder}</b>\n\n"
        f"Endi mijozlar karta orqali to'lov qilganda shu ma'lumotni ko'radi.",
        parse_mode="HTML",
        reply_markup=back_kb(f"shop_settings_{shop_id}")
    )
    return ConversationHandler.END
# ─── ADMIN TELEFON RAQAM BOSHQARUVI ───────────────────────────────────────────
async def add_phone_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id
    if tg_id != ADMIN_ID:
        await q.answer("❌ Ruxsat yo'q!", show_alert=True)
        return ConversationHandler.END

    await q.edit_message_text(
        "📞 <b>Yangi telefon raqam qo'shish</b>\n\n"
        "<b>1-qadam:</b> Ism yoki lavozimni kiriting:\n"
        "<i>Masalan: Operator Sherzod</i>",
        parse_mode="HTML",
        reply_markup=back_kb("admin_settings")
    )
    return WAITING_PHONE_NAME

async def got_admin_phone_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if len(name) < 2:
        await update.message.reply_text("❌ Ism juda qisqa, qaytadan kiriting:")
        return WAITING_PHONE_NAME
    context.user_data["new_phone_name"] = name
    await update.message.reply_text(
        f"✅ Ism: <b>{name}</b>\n\n"
        f"<b>2-qadam:</b> Telefon raqamni kiriting:\n"
        f"<i>Masalan: +998901234567</i>",
        parse_mode="HTML"
    )
    return WAITING_PHONE_NUMBER

async def got_admin_phone_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = update.message.text.strip()
    if not phone.startswith("+") or len(phone) < 9:
        await update.message.reply_text(
            "❌ Noto'g'ri format. Masalan: <code>+998901234567</code>",
            parse_mode="HTML"
        )
        return WAITING_PHONE_NUMBER

    name = context.user_data.pop("new_phone_name", "")
    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []
    phones.append({"name": name, "phone": phone})
    set_setting("order_phones", json.dumps(phones, ensure_ascii=False))

    await update.message.reply_text(
        f"✅ <b>Telefon raqam qo'shildi!</b>\n\n"
        f"👤 Ism: <b>{name}</b>\n"
        f"📱 Raqam: <code>{phone}</code>",
        parse_mode="HTML",
        reply_markup=back_kb("admin_settings")
    )
    return ConversationHandler.END

async def del_phone_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id
    if tg_id != ADMIN_ID:
        await q.answer("❌ Ruxsat yo'q!", show_alert=True)
        return

    idx = int(q.data.split("_")[2])
    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []

    if 0 <= idx < len(phones):
        removed = phones.pop(idx)
        set_setting("order_phones", json.dumps(phones, ensure_ascii=False))
        await q.answer(f"✅ {removed['name']} o'chirildi!", show_alert=True)
    else:
        await q.answer("❌ Topilmadi!", show_alert=True)

    # Sahifani yangilash
    await admin_settings(update, context)

# ─── DO'KON EGASI TELEFON RAQAM BOSHQARUVI ────────────────────────────────────
async def shop_phone_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])
    context.user_data["shop_phone_shop_id"] = shop_id

    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []

    conn = get_db()
    shop = conn.execute("SELECT name FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()
    shop_name = shop["name"] if shop else ""

    # Bu do'konga tegishli raqamlar
    my_phones = [p for p in phones if p.get("shop_id") == shop_id]
    my_phones_text = ""
    if my_phones:
        my_phones_text = "\n\n📞 <b>Joriy raqamlaringiz:</b>\n"
        for p in my_phones:
            my_phones_text += f"👤 {p['name']}: <code>{p['phone']}</code>\n"

    await q.edit_message_text(
        f"📞 <b>Telefon raqam qo'shish</b>{my_phones_text}\n\n"
        f"<b>1-qadam:</b> Mas'ul shaxs ismini kiriting:\n"
        f"<i>Masalan: Operator Sherzod</i>",
        parse_mode="HTML",
        reply_markup=back_kb(f"shop_settings_{shop_id}")
    )
    return WAITING_PHONE_NAME

async def got_shop_phone_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if len(name) < 2:
        await update.message.reply_text("❌ Ism juda qisqa, qaytadan kiriting:")
        return WAITING_PHONE_NAME
    context.user_data["new_shop_phone_name"] = name
    await update.message.reply_text(
        f"✅ Ism: <b>{name}</b>\n\n"
        f"<b>2-qadam:</b> Telefon raqamni kiriting:\n"
        f"<i>Masalan: +998901234567</i>",
        parse_mode="HTML"
    )
    return WAITING_PHONE_NUMBER

async def got_shop_phone_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = update.message.text.strip()
    if not phone.startswith("+") or len(phone) < 9:
        await update.message.reply_text(
            "❌ Noto'g'ri format. Masalan: <code>+998901234567</code>",
            parse_mode="HTML"
        )
        return WAITING_PHONE_NUMBER

    name = context.user_data.pop("new_shop_phone_name", "")
    shop_id = context.user_data.pop("shop_phone_shop_id", None)

    conn = get_db()
    shop = conn.execute("SELECT name FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()
    shop_name = shop["name"] if shop else ""

    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []
    phones.append({"name": name, "phone": phone, "shop_id": shop_id, "shop_name": shop_name})
    set_setting("order_phones", json.dumps(phones, ensure_ascii=False))

    await update.message.reply_text(
        f"✅ <b>Telefon raqam qo'shildi!</b>\n\n"
        f"👤 Ism: <b>{name}</b>\n"
        f"📱 Raqam: <code>{phone}</code>\n"
        f"🏪 Do'kon: <b>{shop_name}</b>\n\n"
        f"Endi mijozlar 'Telefon orqali buyurtma' bo'limida shu raqamni ko'radi.",
        parse_mode="HTML",
        reply_markup=back_kb(f"shop_settings_{shop_id}")
    )
    return ConversationHandler.END

async def shop_del_phone_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    # shop_del_phone_{shop_id}_{idx}
    parts = q.data.split("_")
    shop_id = int(parts[3])
    idx = int(parts[4])

    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []

    # Faqat shu do'konga tegishli raqamlar indeksi
    shop_phones_idx = [i for i, p in enumerate(phones) if p.get("shop_id") == shop_id]

    if 0 <= idx < len(shop_phones_idx):
        real_idx = shop_phones_idx[idx]
        removed = phones.pop(real_idx)
        set_setting("order_phones", json.dumps(phones, ensure_ascii=False))
        await q.answer(f"✅ {removed['name']} o'chirildi!", show_alert=True)
    else:
        await q.answer("❌ Topilmadi!", show_alert=True)
        return

    # shop_settings sahifasiga qaytish
    context_data_backup = q.data
    q.data = f"shop_settings_{shop_id}"
    await shop_settings(update, context)
    q.data = context_data_backup

async def job_apply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🏪 Do'kon egasi bo'lish", callback_data="job_shop_owner")],
        [InlineKeyboardButton("🚴 Kuryer bo'lish", callback_data="job_courier")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")],
    ])
    await q.edit_message_text(
        "💼 <b>Ishga kirish</b>\n\n"
        "Qaysi lavozimga ariza topshirmoqchisiz?\n\n"
        "🏪 <b>Do'kon egasi</b> — o'z do'koningizni oching, mahsulot soting\n"
        "🚴 <b>Kuryer</b> — buyurtmalarni yetkazing va daromad toping",
        reply_markup=kb,
        parse_mode="HTML"
    )

async def job_shop_owner_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    existing = conn.execute("SELECT id FROM shops WHERE owner_tg_id=?", (tg_id,)).fetchone()
    conn.close()

    if existing:
        await q.edit_message_text(
            "❗ Sizda allaqachon do'kon mavjud yoki so'rov yuborilgan.",
            reply_markup=back_kb("job_apply")
        )
        return ConversationHandler.END

    await q.edit_message_text(
        "🏪 <b>Do'kon egasi bo'lish uchun ariza</b>\n\n"
        "Do'koningiz nomini kiriting:",
        parse_mode="HTML"
    )
    return WAITING_JOB_SHOP_NAME

async def got_job_shop_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["job_shop_name"] = update.message.text.strip()
    await update.message.reply_text("📝 Do'kon haqida qisqacha tavsif yozing:")
    return WAITING_JOB_SHOP_DESC

async def got_job_shop_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = context.user_data.pop("job_shop_name", "")
    desc = update.message.text.strip()
    tg_id = update.effective_user.id
    full_name = update.effective_user.full_name or ""

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO shops (owner_tg_id, name, description, status) VALUES (?, ?, ?, 'pending')",
        (tg_id, name, desc)
    )
    shop_id = cur.lastrowid
    conn.commit()
    conn.close()

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"admin_shop_approve_{shop_id}"),
         InlineKeyboardButton("❌ Rad etish", callback_data=f"admin_shop_reject_{shop_id}")],
    ])
    try:
        await update.get_bot().send_message(
            ADMIN_ID,
            f"📋 <b>Yangi do'kon egasi arizasi</b>\n\n"
            f"👤 Ariza beruvchi: {full_name} (ID: {tg_id})\n"
            f"🏪 Do'kon nomi: {name}\n"
            f"📝 Tavsif: {desc}",
            reply_markup=kb,
            parse_mode="HTML"
        )
    except:
        pass

    await update.message.reply_text(
        f"✅ <b>Arizangiz qabul qilindi!</b>\n\n"
        f"🏪 Do'kon nomi: {name}\n"
        f"⏳ Admin ko'rib chiqadi va javob beradi.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Bosh sahifa", callback_data="main_menu")]])
    )
    return ConversationHandler.END

async def job_courier_apply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id
    full_name = update.effective_user.full_name or ""

    conn = get_db()
    existing = conn.execute("SELECT id FROM couriers WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()

    if existing:
        await q.edit_message_text(
            "❗ Siz allaqachon kuryer sifatida ro'yxatdasiz.",
            reply_markup=back_kb()
        )
        return

    try:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Qabul qilish", callback_data=f"admin_courier_approve_{tg_id}"),
             InlineKeyboardButton("❌ Rad etish", callback_data=f"admin_courier_reject_{tg_id}")],
        ])
        await q.get_bot().send_message(
            ADMIN_ID,
            f"🚴 <b>Yangi kuryer arizasi</b>\n\n"
            f"👤 Ism: {full_name}\n"
            f"🆔 Telegram ID: {tg_id}",
            reply_markup=kb,
            parse_mode="HTML"
        )
    except:
        pass

    await q.edit_message_text(
        f"✅ <b>Kuryer arizangiz yuborildi!</b>\n\n"
        f"⏳ Admin ko'rib chiqadi va javob beradi.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Bosh sahifa", callback_data="main_menu")]])
    )

async def admin_courier_approve(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    courier_tg_id = int(q.data.split("_")[3])

    conn = get_db()
    user = conn.execute("SELECT full_name FROM users WHERE tg_id=?", (courier_tg_id,)).fetchone()
    name = user["full_name"] if user else f"Kuryer {courier_tg_id}"
    try:
        conn.execute(
            "INSERT OR IGNORE INTO couriers (tg_id, name, is_premium) VALUES (?, ?, 0)",
            (courier_tg_id, name)
        )
        conn.execute("UPDATE users SET role='courier' WHERE tg_id=?", (courier_tg_id,))
        conn.commit()
    except:
        pass
    conn.close()

    await q.edit_message_text(f"✅ Kuryer {name} qabul qilindi!")
    try:
        await context.bot.send_message(
            courier_tg_id,
            "🎉 <b>Tabriklaymiz!</b>\nSiz OsonSavdo kuryeri sifatida qabul qilindingiz!\n/start bosing.",
            parse_mode="HTML"
        )
    except:
        pass

async def admin_courier_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    courier_tg_id = int(q.data.split("_")[3])

    await q.edit_message_text(f"❌ Kuryer arizasi rad etildi.")
    try:
        await context.bot.send_message(
            courier_tg_id,
            "❌ Afsuski, kuryer arizangiz rad etildi."
        )
    except:
        pass

# ─── ADMIN DO'KON QO'SHISH ─────────────────────────────────────────────────────
async def admin_add_shop_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    if update.effective_user.id != ADMIN_ID:
        await q.answer("❌ Ruxsat yo'q!", show_alert=True)
        return ConversationHandler.END
    await q.edit_message_text(
        "🏪 <b>Admin: Yangi do'kon qo'shish</b>\n\n"
        "Do'kon egasining Telegram ID sini kiriting:",
        parse_mode="HTML"
    )
    return WAITING_ADMIN_SHOP_OWNER

async def got_admin_shop_owner(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        owner_id = int(text)
        if owner_id <= 0:
            raise ValueError
        context.user_data["admin_shop_owner_id"] = owner_id
        await update.message.reply_text("🏪 Do'kon nomini kiriting:")
        return WAITING_ADMIN_SHOP_NAME
    except:
        await update.message.reply_text(
            f"❌ Noto'g'ri Telegram ID.\n"
            f"Faqat musbat raqam kiriting.\nMasalan: 123456789"
        )
        return WAITING_ADMIN_SHOP_OWNER

async def got_admin_shop_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["admin_shop_name"] = update.message.text.strip()
    await update.message.reply_text("📝 Do'kon tavsifini kiriting:")
    return WAITING_ADMIN_SHOP_DESC

async def got_admin_shop_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    owner_id = context.user_data.pop("admin_shop_owner_id", None)
    name = context.user_data.pop("admin_shop_name", "")
    desc = update.message.text.strip()

    if not owner_id:
        await update.message.reply_text("❌ Xatolik. Qaytadan urining.")
        return ConversationHandler.END

    conn = get_db()
    conn.execute(
        "INSERT INTO shops (owner_tg_id, name, description, status) VALUES (?, ?, ?, 'approved')",
        (owner_id, name, desc)
    )
    conn.execute("INSERT OR IGNORE INTO users (tg_id, full_name, role) VALUES (?, ?, 'shop_owner')", (owner_id, name))
    conn.execute("UPDATE users SET role='shop_owner' WHERE tg_id=?", (owner_id,))
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"✅ Do'kon <b>{name}</b> muvaffaqiyatli qo'shildi va tasdiqlandi!",
        parse_mode="HTML",
        reply_markup=back_kb("admin_panel")
    )
    try:
        await context.bot.send_message(
            owner_id,
            f"🎉 <b>Tabriklaymiz!</b>\nSizning <b>{name}</b> do'koningiz admin tomonidan qo'shildi!\n/start bosing.",
            parse_mode="HTML"
        )
    except:
        pass
    return ConversationHandler.END

async def admin_job_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    pending = conn.execute(
        "SELECT s.*, u.full_name FROM shops s LEFT JOIN users u ON s.owner_tg_id=u.tg_id WHERE s.status='pending' ORDER BY s.id DESC"
    ).fetchall()
    conn.close()

    if not pending:
        await q.edit_message_text("📋 Kutayotgan arizalar yo'q.", reply_markup=back_kb("admin_panel"))
        return

    buttons = []
    for s in pending:
        buttons.append([InlineKeyboardButton(
            f"⏳ {s['name']} — {s['full_name'] or s['owner_tg_id']}",
            callback_data=f"admin_shop_{s['id']}"
        )])
    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")])

    await q.edit_message_text(
        "📋 <b>Ishga kirish arizalari (do'kon):</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

# ─── ADD SHOP ──────────────────────────────────────────────────────────────────
async def add_shop_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data["adding_shop"] = True
    await q.edit_message_text("🏪 Do'kon nomini kiriting:")
    return WAITING_SHOP_NAME

async def got_shop_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["shop_name"] = update.message.text.strip()
    await update.message.reply_text("📝 Do'kon tavsifini kiriting:")
    return WAITING_SHOP_DESC

async def got_shop_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = context.user_data.pop("shop_name", "")
    desc = update.message.text.strip()
    tg_id = update.effective_user.id

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO shops (owner_tg_id, name, description) VALUES (?, ?, ?)",
        (tg_id, name, desc)
    )
    shop_id = cur.lastrowid

    # Set user as shop_owner
    conn.execute("UPDATE users SET role='shop_owner' WHERE tg_id=?", (tg_id,))
    conn.commit()
    conn.close()

    # Notify admin
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"admin_shop_approve_{shop_id}"),
         InlineKeyboardButton("❌ Rad etish", callback_data=f"admin_shop_reject_{shop_id}")],
    ])
    try:
        await context.bot.send_message(
            ADMIN_ID,
            f"🏪 <b>Yangi do'kon so'rovi</b>\n\n"
            f"Nomi: {name}\nTavsif: {desc}\n"
            f"Egasi ID: {tg_id}",
            reply_markup=kb,
            parse_mode="HTML"
        )
    except:
        pass

    await update.message.reply_text(
        "✅ Do'kon qo'shish so'rovi yuborildi!\nAdmin tasdiqlashini kuting.",
        reply_markup=back_kb()
    )
    return ConversationHandler.END

# ─── ADD PRODUCT ───────────────────────────────────────────────────────────────
async def add_product_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    shop = conn.execute("SELECT id FROM shops WHERE owner_tg_id=? AND status='approved'", (tg_id,)).fetchone()
    conn.close()

    if not shop:
        await q.edit_message_text("❌ Tasdiqlangan do'koningiz yo'q.", reply_markup=back_kb())
        return ConversationHandler.END

    context.user_data["adding_product_shop"] = shop["id"]
    await q.edit_message_text("🛍 Mahsulot nomini kiriting:")
    return WAITING_PRODUCT_NAME

async def got_product_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod_name"] = update.message.text.strip()
    await update.message.reply_text("📝 Mahsulot tavsifini kiriting:")
    return WAITING_PRODUCT_DESC

async def got_product_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod_desc"] = update.message.text.strip()
    await update.message.reply_text("💰 Narxini kiriting (so'mda):")
    return WAITING_PRODUCT_PRICE

async def got_product_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        price = float(update.message.text.strip().replace(" ", "").replace(",", ""))
        context.user_data["prod_price"] = price
        await update.message.reply_text("📸 Rasmini yuboring (yoki /skip):")
        return WAITING_PRODUCT_PHOTO
    except:
        await update.message.reply_text("❌ Noto'g'ri format. Faqat raqam kiriting:")
        return WAITING_PRODUCT_PRICE

async def got_product_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text and update.message.text == "/skip":
        photo_id = None
    elif update.message.photo:
        photo_id = update.message.photo[-1].file_id
    else:
        photo_id = None

    shop_id = context.user_data.pop("adding_product_shop", None)
    name = context.user_data.pop("prod_name", "")
    desc = context.user_data.pop("prod_desc", "")
    price = context.user_data.pop("prod_price", 0)

    conn = get_db()
    conn.execute(
        "INSERT INTO products (shop_id, name, description, price, photo_id) VALUES (?, ?, ?, ?, ?)",
        (shop_id, name, desc, price, photo_id)
    )
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"✅ <b>{name}</b> mahsuloti qo'shildi!\n💰 Narx: {format_price(price)}",
        reply_markup=back_kb("my_shop"),
        parse_mode="HTML"
    )
    return ConversationHandler.END

# ─── OWNER PRODUCTS ────────────────────────────────────────────────────────────
async def owner_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])

    conn = get_db()
    products = conn.execute("SELECT * FROM products WHERE shop_id=?", (shop_id,)).fetchall()
    conn.close()

    if not products:
        await q.edit_message_text("📦 Mahsulotlar yo'q.", reply_markup=back_kb("my_shop"))
        return

    buttons = []
    for p in products:
        status = "✅" if p["is_active"] else "❌"
        buttons.append([InlineKeyboardButton(
            f"{status} {p['name']} — {format_price(p['price'])}",
            callback_data=f"owner_prod_{p['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="my_shop")])
    await q.edit_message_text(
        "📦 <b>Mahsulotlaringiz:</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def owner_product_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    prod_id = int(q.data.split("_")[2])

    conn = get_db()
    p = conn.execute("SELECT * FROM products WHERE id=?", (prod_id,)).fetchone()
    conn.close()

    if not p:
        await q.answer("Topilmadi", show_alert=True)
        return

    toggle_label = "❌ O'chirish" if p["is_active"] else "✅ Yoqish"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✏️ Narx o'zgartiris", callback_data=f"edit_price_{prod_id}"),
         InlineKeyboardButton("🔥 Aksiya", callback_data=f"set_discount_{prod_id}")],
        [InlineKeyboardButton(toggle_label, callback_data=f"toggle_prod_{prod_id}")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data=f"owner_products_{p['shop_id']}")],
    ])

    disc_txt = f"\n🔥 Aksiya: {p['discount_percent']}%" if p["discount_percent"] else ""
    await q.edit_message_text(
        f"🛍 <b>{p['name']}</b>\n"
        f"💰 {format_price(p['price'])}{disc_txt}\n"
        f"📦 Ombor: {p['stock']}\n"
        f"{'✅ Faol' if p['is_active'] else '❌ Nofaol'}",
        reply_markup=kb,
        parse_mode="HTML"
    )

async def edit_product_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    prod_id = int(q.data.split("_")[2])
    context.user_data["editing_prod_id"] = prod_id
    await q.edit_message_text("💰 Yangi narxni kiriting (so'mda):")
    return WAITING_PRODUCT_EDIT_PRICE

async def got_edit_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        price = float(update.message.text.strip().replace(" ", ""))
    except:
        await update.message.reply_text("❌ Noto'g'ri format.")
        return WAITING_PRODUCT_EDIT_PRICE

    prod_id = context.user_data.pop("editing_prod_id", None)
    if prod_id:
        conn = get_db()
        conn.execute("UPDATE products SET price=? WHERE id=?", (price, prod_id))
        conn.commit()
        conn.close()

    await update.message.reply_text(f"✅ Narx {format_price(price)} ga o'zgartirildi!", reply_markup=back_kb("my_shop"))
    return ConversationHandler.END

async def set_product_discount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    prod_id = int(q.data.split("_")[2])
    context.user_data["discounting_prod_id"] = prod_id
    await q.edit_message_text("🔥 Chegirma foizini kiriting (0-100):")
    return WAITING_PRODUCT_DISCOUNT

async def got_product_discount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        pct = float(update.message.text.strip())
        if not 0 <= pct <= 100:
            raise ValueError
    except:
        await update.message.reply_text("❌ 0 dan 100 gacha raqam kiriting.")
        return WAITING_PRODUCT_DISCOUNT

    prod_id = context.user_data.pop("discounting_prod_id", None)
    if prod_id:
        conn = get_db()
        conn.execute("UPDATE products SET discount_percent=? WHERE id=?", (pct, prod_id))
        conn.commit()
        conn.close()

    await update.message.reply_text(f"✅ Aksiya {int(pct)}% o'rnatildi!", reply_markup=back_kb("my_shop"))
    return ConversationHandler.END

async def toggle_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    prod_id = int(q.data.split("_")[2])

    conn = get_db()
    p = conn.execute("SELECT is_active, shop_id FROM products WHERE id=?", (prod_id,)).fetchone()
    if p:
        new_status = 0 if p["is_active"] else 1
        conn.execute("UPDATE products SET is_active=? WHERE id=?", (new_status, prod_id))
        conn.commit()
    conn.close()

    await q.answer("✅ O'zgartirildi!")
    await owner_product_detail(update, context)

# ─── SHOP ORDERS ───────────────────────────────────────────────────────────────
async def shop_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    shop = conn.execute("SELECT id FROM shops WHERE owner_tg_id=?", (tg_id,)).fetchone()
    if not shop:
        await q.edit_message_text("Do'kon topilmadi.", reply_markup=back_kb())
        conn.close()
        return

    orders = conn.execute(
        "SELECT * FROM orders WHERE shop_id=? ORDER BY id DESC LIMIT 20",
        (shop["id"],)
    ).fetchall()
    conn.close()

    if not orders:
        await q.edit_message_text("📦 Buyurtmalar yo'q.", reply_markup=back_kb())
        return

    buttons = []
    for o in orders:
        emoji = order_status_emoji(o["status"])
        buttons.append([InlineKeyboardButton(
            f"{emoji} #{1000 + o['id']} — {format_price(o['total'])} — {order_status_text(o['status'])}",
            callback_data=f"shop_order_detail_{o['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="my_shop")])
    await q.edit_message_text(
        "📦 <b>Do'kon buyurtmalari</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

# ─── SHOP REPORT ───────────────────────────────────────────────────────────────
async def shop_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE owner_tg_id=?", (tg_id,)).fetchone()
    if not shop:
        await q.edit_message_text("Do'kon topilmadi.", reply_markup=back_kb())
        conn.close()
        return

    thirty_days_ago = (datetime.now() - timedelta(days=30)).isoformat()
    orders = conn.execute(
        "SELECT * FROM orders WHERE shop_id=? AND created_at>? AND status='delivered'",
        (shop["id"], thirty_days_ago)
    ).fetchall()
    conn.close()

    total_revenue = sum(o["total"] - o["commission"] for o in orders)
    total_orders = len(orders)

    if not EXCEL_AVAILABLE:
        await q.edit_message_text(
            f"📊 <b>30 kunlik hisobot</b>\n\n"
            f"📦 Buyurtmalar: {total_orders}\n"
            f"💰 Daromad: {format_price(total_revenue)}\n\n"
            f"Excel uchun openpyxl o'rnating.",
            reply_markup=back_kb("my_shop"),
            parse_mode="HTML"
        )
        return

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="4CAF50")

    headers = ["#", "Buyurtma ID", "Sana", "Jami", "Komissiya", "Daromad", "Holat"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, o in enumerate(orders, 2):
        revenue = o["total"] - o["commission"]
        ws.append([
            row - 1,
            f"#{1000 + o['id']}",
            o["created_at"][:10],
            o["total"],
            o["commission"],
            revenue,
            order_status_text(o["status"])
        ])

    # Summary
    ws.append([])
    ws.append(["", "", "JAMI:", total_revenue + sum(o["commission"] for o in orders),
               sum(o["commission"] for o in orders), total_revenue, ""])

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    await context.bot.send_document(
        tg_id,
        document=InputFile(excel_buf, filename=f"hisobot_{shop['name']}_{datetime.now().strftime('%Y%m%d')}.xlsx"),
        caption=f"📊 <b>30 kunlik hisobot</b>\n📦 Buyurtmalar: {total_orders}\n💰 Daromad: {format_price(total_revenue)}",
        parse_mode="HTML"
    )

# ─── ADMIN PANEL ───────────────────────────────────────────────────────────────
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    if update.effective_user.id != ADMIN_ID:
        await q.answer("❌ Ruxsat yo'q!", show_alert=True)
        return

    conn = get_db()
    total_users = conn.execute("SELECT COUNT(*) as c FROM users").fetchone()["c"]
    total_orders = conn.execute("SELECT COUNT(*) as c FROM orders").fetchone()["c"]
    total_shops = conn.execute("SELECT COUNT(*) as c FROM shops WHERE status='approved'").fetchone()["c"]
    total_revenue = conn.execute("SELECT SUM(commission) as s FROM orders WHERE status='delivered'").fetchone()["s"] or 0
    pending_shops = conn.execute("SELECT COUNT(*) as c FROM shops WHERE status='pending'").fetchone()["c"]
    open_tickets = conn.execute("SELECT COUNT(*) as c FROM tickets WHERE status='open'").fetchone()["c"]
    conn.close()

    commission = get_setting("commission_percent", "10")

    text = (
        f"⚙️ <b>Admin Panel</b>\n\n"
        f"👥 Foydalanuvchilar: {total_users}\n"
        f"📦 Buyurtmalar: {total_orders}\n"
        f"🏪 Do'konlar: {total_shops}\n"
        f"💰 Platforma daromadi: {format_price(total_revenue)}\n"
        f"📊 Komissiya: {commission}%\n"
        f"⏳ Kutayotgan do'konlar: {pending_shops}\n"
        f"🎟 Ochiq ticketlar: {open_tickets}\n"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🏪 Do'konlar", callback_data="admin_shops"),
         InlineKeyboardButton("👥 Foydalanuvchilar", callback_data="admin_users")],
        [InlineKeyboardButton("📦 Buyurtmalar", callback_data="admin_orders"),
         InlineKeyboardButton("🚴 Kuryerlar", callback_data="admin_couriers")],
        [InlineKeyboardButton("🎫 Promo kodlar", callback_data="admin_promos"),
         InlineKeyboardButton("🎟 Ticketlar", callback_data="admin_tickets")],
        [InlineKeyboardButton("📊 Excel hisobot", callback_data="admin_excel"),
         InlineKeyboardButton("💰 Komissiya", callback_data="admin_commission")],
        [InlineKeyboardButton("📢 Broadcast", callback_data="admin_broadcast"),
         InlineKeyboardButton("⚙️ Sozlamalar", callback_data="admin_settings")],
        [InlineKeyboardButton("➕ Do'kon qo'shish", callback_data="admin_add_shop"),
         InlineKeyboardButton("📋 Ishga kirish arizalar", callback_data="admin_job_requests")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")],
    ])
    await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

async def admin_shops(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    shops = conn.execute("SELECT * FROM shops ORDER BY id DESC").fetchall()
    conn.close()

    buttons = []
    for s in shops:
        status = {"approved": "✅", "pending": "⏳", "rejected": "❌"}.get(s["status"], "?")
        buttons.append([InlineKeyboardButton(
            f"{status} {s['name']}", callback_data=f"admin_shop_{s['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")])
    await q.edit_message_text(
        "🏪 <b>Barcha do'konlar:</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def admin_shop_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])

    conn = get_db()
    s = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()

    if not s:
        await q.answer("Topilmadi", show_alert=True)
        return

    buttons = []
    if s["status"] == "pending":
        buttons.append([
            InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"admin_shop_approve_{shop_id}"),
            InlineKeyboardButton("❌ Rad etish", callback_data=f"admin_shop_reject_{shop_id}"),
        ])
    elif s["status"] == "approved":
        buttons.append([InlineKeyboardButton("🚫 Bloklash", callback_data=f"admin_shop_reject_{shop_id}")])

    is_open = s["is_open"] if "is_open" in s.keys() else 1
    open_txt = "🟢 Ochiq" if is_open else "🔴 Yopiq"
    admin_toggle_label = "🔴 Yopish" if is_open else "🟢 Ochish"
    buttons.append([InlineKeyboardButton(f"{admin_toggle_label} (admin)", callback_data=f"admin_toggle_shop_{shop_id}")])
    buttons.append([InlineKeyboardButton("💰 Obuna % sozlash", callback_data=f"admin_sub_{shop_id}")])
    buttons.append([InlineKeyboardButton("💳 To'lov tizimini tahrirlash", callback_data=f"admin_set_payment_{shop_id}")])
    buttons.append([InlineKeyboardButton("🗑 Do'konni o'chirish", callback_data=f"admin_shop_delete_{shop_id}")])
    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_shops")])

    card_text = ""
    if s["card_number"]:
        cn = s["card_number"]
        formatted_card = " ".join([cn[i:i+4] for i in range(0, len(cn), 4)])
        card_text = f"\n💳 Karta: <b>{formatted_card}</b>\n👤 Ism: <b>{s['card_holder'] or 'Kiritilmagan'}</b>"
    else:
        card_text = "\n💳 Karta: <i>Kiritilmagan</i>"

    # Obuna ma'lumoti
    conn2 = get_db()
    sub = conn2.execute("SELECT * FROM shop_subscriptions WHERE shop_id=?", (shop_id,)).fetchone()
    conn2.close()
    if sub:
        due = sub["next_due_at"][:10] if sub["next_due_at"] else "—"
        earned = format_price(sub["total_earned"])
        sub_text = (
            f"\n\n📅 <b>Obuna:</b> {sub['fee_percent']}% har 30 kun\n"
            f"🗓 Keyingi to'lov: <b>{due}</b>\n"
            f"💵 Jami tushgan: <b>{earned}</b>"
        )
    else:
        sub_text = "\n\n📅 <b>Obuna:</b> <i>Belgilanmagan</i>"

    await q.edit_message_text(
        f"🏪 <b>{s['name']}</b>\n"
        f"📝 {s['description']}\n"
        f"👤 Egasi ID: {s['owner_tg_id']}\n"
        f"📊 Holat: {s['status']}\n"
        f"🚦 Do'kon: {open_txt}\n"
        f"⭐ {s['rating']:.1f} ({s['total_reviews']} sharh)"
        f"{card_text}"
        f"{sub_text}",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def admin_shop_approve(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[3])

    conn = get_db()
    s = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.execute("UPDATE shops SET status='approved' WHERE id=?", (shop_id,))
    conn.commit()
    conn.close()

    await q.edit_message_text(f"✅ Do'kon #{shop_id} tasdiqlandi!")
    if s:
        try:
            await context.bot.send_message(s["owner_tg_id"], f"🎉 Do'koningiz '{s['name']}' tasdiqlandi! /start bosing.")
        except:
            pass

async def admin_shop_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[3])

    conn = get_db()
    s = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.execute("UPDATE shops SET status='rejected' WHERE id=?", (shop_id,))
    conn.commit()
    conn.close()

    await q.edit_message_text(f"❌ Do'kon #{shop_id} rad etildi.")
    if s:
        try:
            await context.bot.send_message(s["owner_tg_id"], f"❌ Do'koningiz '{s['name']}' rad etildi.")
        except:
            pass

async def admin_shop_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[3])

    conn = get_db()
    s = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    if s:
        conn.execute("DELETE FROM products WHERE shop_id=?", (shop_id,))
        conn.execute("DELETE FROM favorites WHERE shop_id=?", (shop_id,))
        conn.execute("DELETE FROM reviews WHERE shop_id=?", (shop_id,))
        conn.execute("UPDATE users SET role='customer' WHERE tg_id=?", (s["owner_tg_id"],))
        conn.execute("DELETE FROM shops WHERE id=?", (shop_id,))
        conn.commit()
        try:
            await context.bot.send_message(
                s["owner_tg_id"],
                f"❌ Sizning <b>{s['name']}</b> do'koningiz admin tomonidan o'chirildi.",
                parse_mode="HTML"
            )
        except:
            pass
    conn.close()

    await q.edit_message_text(
        f"🗑 Do'kon muvaffaqiyatli o'chirildi.",
        reply_markup=back_kb("admin_shops")
    )

async def admin_toggle_shop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[3])

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    if not shop:
        conn.close()
        return

    new_state = 0 if (shop["is_open"] if "is_open" in shop.keys() else 1) else 1
    conn.execute("UPDATE shops SET is_open=? WHERE id=?", (new_state, shop_id))
    conn.commit()

    # Do'kon egasini xabardor qil
    state_txt = "🟢 Ochiq" if new_state else "🔴 Yopiq"
    try:
        await context.bot.send_message(
            shop["owner_tg_id"],
            f"⚠️ <b>Do'koningiz holati o'zgartirildi!</b>\n\n"
            f"🏪 {shop['name']}\n"
            f"🚦 Yangi holat: <b>{state_txt}</b>\n\n"
            f"{'Endi mijozlar zakaz bera oladi.' if new_state else 'Mijozlar hozir zakaz bera olmaydi.'}",
            parse_mode="HTML"
        )
    except:
        pass
    conn.close()

    await q.answer(f"Do'kon {state_txt} qilindi!", show_alert=True)
    # Sahifani yangilash
    q.data = f"admin_shop_{shop_id}"
    await admin_shop_detail(update, context)

# ─── ADMIN OBUNA TIZIMI (30 KUNLIK %) ─────────────────────────────────────────
async def admin_sub_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    shop_id = int(q.data.split("_")[2])
    context.user_data["sub_shop_id"] = shop_id

    conn = get_db()
    shop = conn.execute("SELECT name FROM shops WHERE id=?", (shop_id,)).fetchone()
    sub = conn.execute("SELECT * FROM shop_subscriptions WHERE shop_id=?", (shop_id,)).fetchone()
    conn.close()

    if sub:
        due = sub["next_due_at"][:10] if sub["next_due_at"] else "—"
        earned = format_price(sub["total_earned"])
        cur_info = (
            f"\n\n📌 Joriy obuna:\n"
            f"💰 Foiz: <b>{sub['fee_percent']}%</b>\n"
            f"🗓 Keyingi to'lov: <b>{due}</b>\n"
            f"💵 Jami tushgan: <b>{earned}</b>\n"
            f"📊 Holat: {'✅ Faol' if sub['status'] == 'active' else '❌ Nofaol'}"
        )
    else:
        cur_info = "\n\n📌 Obuna hali belgilanmagan."

    await q.edit_message_text(
        f"💰 <b>'{shop['name'] if shop else ''}' obuna sozlamalari</b>{cur_info}\n\n"
        f"Har 30 kunda do'kon daromadidan olinadigan foizni kiriting (0-50):\n"
        f"Masalan: <code>10</code> (10% degani)",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Orqaga", callback_data=f"admin_shop_{shop_id}")]
        ])
    )
    return WAITING_SUB_PERCENT

async def got_sub_percent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        pct = float(update.message.text.strip().replace("%", ""))
        if not 0 <= pct <= 50:
            raise ValueError
    except:
        await update.message.reply_text("❌ 0 dan 50 gacha raqam kiriting. Masalan: <code>10</code>", parse_mode="HTML")
        return WAITING_SUB_PERCENT

    shop_id = context.user_data.pop("sub_shop_id", None)
    if not shop_id:
        return ConversationHandler.END

    now = datetime.now()
    next_due = (now + timedelta(days=30)).isoformat()

    conn = get_db()
    existing = conn.execute("SELECT id FROM shop_subscriptions WHERE shop_id=?", (shop_id,)).fetchone()
    if existing:
        conn.execute(
            "UPDATE shop_subscriptions SET fee_percent=?, next_due_at=?, status='active' WHERE shop_id=?",
            (pct, next_due, shop_id)
        )
    else:
        conn.execute(
            "INSERT INTO shop_subscriptions (shop_id, fee_percent, last_paid_at, next_due_at) VALUES (?, ?, ?, ?)",
            (shop_id, pct, now.isoformat(), next_due)
        )
    shop = conn.execute("SELECT name, owner_tg_id FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"✅ <b>Obuna belgilandi!</b>\n\n"
        f"🏪 Do'kon: <b>{shop['name'] if shop else shop_id}</b>\n"
        f"💰 Foiz: <b>{pct}%</b> har 30 kunda\n"
        f"🗓 Birinchi to'lov: <b>{next_due[:10]}</b>",
        parse_mode="HTML",
        reply_markup=back_kb(f"admin_shop_{shop_id}")
    )

    # Do'kon egasini xabardor qilish
    if shop:
        try:
            await update.get_bot().send_message(
                shop["owner_tg_id"],
                f"📢 <b>Obuna tizimi haqida xabar</b>\n\n"
                f"Do'koningiz <b>{shop['name']}</b> uchun obuna belgilandi:\n"
                f"💰 Har 30 kunda bot orqali tushgan daromaddan <b>{pct}%</b> olinadi.\n"
                f"🗓 Birinchi hisob-kitob: <b>{next_due[:10]}</b>",
                parse_mode="HTML"
            )
        except:
            pass

    return ConversationHandler.END

async def auto_open_close_shops(bot):
    """Har daqiqa ishga tushadi — ish vaqtiga qarab do'konni ochadi/yopadi"""
    import re
    now = datetime.now()
    current_minutes = now.hour * 60 + now.minute

    conn = get_db()
    # NULL bo'lgan is_open larni 1 ga o'zgartir
    conn.execute("UPDATE shops SET is_open=1 WHERE is_open IS NULL")
    conn.commit()

    shops = conn.execute(
        "SELECT id, name, owner_tg_id, work_hours, is_open FROM shops WHERE status='approved'"
    ).fetchall()

    for shop in shops:
        wh = shop["work_hours"]
        if not wh or "-" not in wh:
            continue
        if not re.match(r"^\d{1,2}:\d{2}-\d{1,2}:\d{2}$", wh):
            continue

        try:
            open_t, close_t = wh.split("-")
            oh, om = map(int, open_t.split(":"))
            ch, cm = map(int, close_t.split(":"))
        except:
            continue

        open_min = oh * 60 + om
        close_min = ch * 60 + cm
        is_open = 1 if shop["is_open"] is None else int(shop["is_open"])

        # Tungi vaqt (masalan 22:00-02:00)
        if open_min < close_min:
            should_open = open_min <= current_minutes < close_min
        else:
            should_open = current_minutes >= open_min or current_minutes < close_min

        # Holat o'zgargan bo'lsagina yangilash
        if should_open and not is_open:
            conn.execute("UPDATE shops SET is_open=1 WHERE id=?", (shop["id"],))
            try:
                await bot.send_message(
                    shop["owner_tg_id"],
                    f"🟢 <b>Do'koningiz ochildi!</b>\n🏪 {shop['name']}\n⏰ Ish vaqti: {wh}",
                    parse_mode="HTML"
                )
            except:
                pass
        elif not should_open and is_open:
            conn.execute("UPDATE shops SET is_open=0 WHERE id=?", (shop["id"],))
            try:
                await bot.send_message(
                    shop["owner_tg_id"],
                    f"🔴 <b>Do'koningiz yopildi!</b>\n🏪 {shop['name']}\n⏰ Ish vaqti: {wh}",
                    parse_mode="HTML"
                )
            except:
                pass

    conn.commit()
    conn.close()

async def check_subscriptions(context):
    """Har kuni ishga tushadigan: muddati o'tgan obunalarni tekshiradi"""
    conn = get_db()
    now = datetime.now()
    subs = conn.execute(
        "SELECT ss.*, s.name as shop_name, s.owner_tg_id "
        "FROM shop_subscriptions ss JOIN shops s ON ss.shop_id=s.id "
        "WHERE ss.status='active' AND ss.next_due_at <= ?",
        (now.isoformat(),)
    ).fetchall()

    for sub in subs:
        shop_id = sub["shop_id"]
        pct = sub["fee_percent"]

        # Oxirgi 30 kundagi yetkazilgan buyurtmalar summasi
        last_check = sub["last_paid_at"] or (now - timedelta(days=30)).isoformat()
        orders = conn.execute(
            "SELECT SUM(total) as s FROM orders WHERE shop_id=? AND status='delivered' AND updated_at >= ?",
            (shop_id, last_check)
        ).fetchone()
        total_sales = orders["s"] or 0
        fee_amount = total_sales * pct / 100

        next_due = (now + timedelta(days=30)).isoformat()
        conn.execute(
            "UPDATE shop_subscriptions SET last_paid_at=?, next_due_at=?, total_earned=total_earned+? WHERE shop_id=?",
            (now.isoformat(), next_due, fee_amount, shop_id)
        )
        conn.commit()

        # Adminga xabar
        try:
            await context.bot.send_message(
                ADMIN_ID,
                f"💰 <b>Obuna hisob-kitob</b>\n\n"
                f"🏪 Do'kon: <b>{sub['shop_name']}</b>\n"
                f"📊 30 kunlik savdo: <b>{format_price(total_sales)}</b>\n"
                f"💵 Olingan foiz ({pct}%): <b>{format_price(fee_amount)}</b>\n"
                f"🗓 Keyingi hisob: <b>{next_due[:10]}</b>",
                parse_mode="HTML"
            )
        except:
            pass

        # Do'kon egasiga xabar
        try:
            await context.bot.send_message(
                sub["owner_tg_id"],
                f"📅 <b>Oylik obuna hisob-kitob</b>\n\n"
                f"🏪 Do'kon: <b>{sub['shop_name']}</b>\n"
                f"📊 Oylik savdo: <b>{format_price(total_sales)}</b>\n"
                f"💵 Platform ulushi ({pct}%): <b>{format_price(fee_amount)}</b>\n"
                f"🗓 Keyingi hisob: <b>{next_due[:10]}</b>",
                parse_mode="HTML"
            )
        except:
            pass

    conn.close()

# ─── ADMIN TO'LOV TIZIMI TAHRIRLASH ────────────────────────────────────────────
async def admin_set_payment_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    parts = q.data.split("_")
    shop_id = int(parts[3])
    context.user_data["admin_payment_shop_id"] = shop_id

    conn = get_db()
    shop = conn.execute("SELECT * FROM shops WHERE id=?", (shop_id,)).fetchone()
    conn.close()

    current = ""
    if shop and shop["card_number"]:
        cn = shop["card_number"]
        formatted_card = " ".join([cn[i:i+4] for i in range(0, len(cn), 4)])
        current = (
            f"\n\n📌 Joriy ma'lumot:\n"
            f"💳 Karta: <b>{formatted_card}</b>\n"
            f"👤 Ism: <b>{shop['card_holder'] or 'Kiritilmagan'}</b>"
        )

    await q.edit_message_text(
        f"💳 <b>Admin: '{shop['name'] if shop else ''}' do'koni to'lov tizimi</b>{current}\n\n"
        f"Yangi karta raqamini kiriting (13-19 ta raqam):\n"
        f"Masalan: <code>8600123456789012</code>",
        parse_mode="HTML",
        reply_markup=back_kb(f"admin_shop_{shop_id}")
    )
    return WAITING_CARD_NUMBER

async def admin_got_card_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    number = update.message.text.strip().replace(" ", "").replace("-", "")
    if not number.isdigit() or len(number) < 13 or len(number) > 19:
        await update.message.reply_text(
            "❌ Noto'g'ri karta raqami. 13-19 ta raqam kiriting.\n"
            "Masalan: <code>8600123456789012</code>",
            parse_mode="HTML"
        )
        return WAITING_CARD_NUMBER

    context.user_data["admin_new_card_number"] = number
    await update.message.reply_text(
        f"✅ Karta raqami: <b>{number}</b>\n\n"
        f"👤 Karta egasining ismini kiriting:\n"
        f"Masalan: <code>ALISHER KARIMOV</code>",
        parse_mode="HTML"
    )
    return WAITING_CARD_HOLDER

async def admin_got_card_holder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    holder = update.message.text.strip().upper()
    if len(holder) < 3:
        await update.message.reply_text("❌ Ism juda qisqa.")
        return WAITING_CARD_HOLDER

    shop_id = context.user_data.pop("admin_payment_shop_id", None)
    card_number = context.user_data.pop("admin_new_card_number", "")

    if shop_id:
        conn = get_db()
        conn.execute(
            "UPDATE shops SET card_number=?, card_holder=? WHERE id=?",
            (card_number, holder, shop_id)
        )
        conn.commit()
        conn.close()

    formatted = " ".join([card_number[i:i+4] for i in range(0, len(card_number), 4)])
    await update.message.reply_text(
        f"✅ <b>Do'kon to'lov tizimi yangilandi!</b>\n\n"
        f"💳 Karta: <b>{formatted}</b>\n"
        f"👤 Ism: <b>{holder}</b>",
        parse_mode="HTML",
        reply_markup=back_kb(f"admin_shop_{shop_id}")
    )
    return ConversationHandler.END

async def admin_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    users = conn.execute("SELECT * FROM users ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()

    text = "👥 <b>Foydalanuvchilar (so'nggi 20):</b>\n\n"
    for u in users:
        role = u["role"] or "customer"
        text += f"• {u['full_name']} [{role}] — {u['total_orders']} buyurtma\n"

    await q.edit_message_text(text, reply_markup=back_kb("admin_panel"), parse_mode="HTML")

async def admin_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    orders = conn.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id ORDER BY o.id DESC LIMIT 15"
    ).fetchall()
    conn.close()

    buttons = []
    for o in orders:
        emoji = order_status_emoji(o["status"])
        buttons.append([InlineKeyboardButton(
            f"{emoji} #{1000 + o['id']} | {o['shop_name']} | {format_price(o['total'])}",
            callback_data=f"admin_order_{o['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")])
    await q.edit_message_text(
        "📦 <b>Barcha buyurtmalar:</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def admin_order_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    order_id = int(q.data.split("_")[2])

    conn = get_db()
    o = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    conn.close()

    if not o:
        await q.answer("Topilmadi", show_alert=True)
        return

    items = json.loads(o["items"])
    items_txt = "\n".join([f"• {v['name']} x{v['qty']}" for v in items.values()])

    text = (
        f"📦 <b>Buyurtma #{1000 + order_id}</b>\n\n"
        f"{items_txt}\n\n"
        f"📍 Manzil: {o['address']}\n"
        f"💰 Jami: {format_price(o['total'])}\n"
        f"📊 Komissiya: {format_price(o['commission'])}\n"
        f"💳 To'lov: {'Karta' if o['payment_method'] == 'card' else 'Naqd'}\n"
        f"🚴 Kuryer: {o['courier_type']}\n"
        f"📊 Holat: {order_status_text(o['status'])}\n"
    )

    buttons = []
    if o["status"] == "new":
        buttons.append([
            InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"admin_confirm_{order_id}"),
            InlineKeyboardButton("❌ Rad etish", callback_data=f"admin_reject_{order_id}"),
        ])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_orders")])
    await q.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode="HTML")

# ─── ADMIN COURIERS ────────────────────────────────────────────────────────────
async def admin_couriers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    couriers = conn.execute("SELECT * FROM couriers ORDER BY id DESC").fetchall()
    conn.close()

    if not couriers:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ Kuryer qo'shish", callback_data="add_courier")],
            [InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")],
        ])
        await q.edit_message_text("🚴 Kuryerlar yo'q.", reply_markup=kb)
        return

    buttons = []
    for c in couriers:
        busy = "🔴" if c["is_busy"] else "🟢"
        prem = "⚡" if c["is_premium"] else "🚴"
        buttons.append([InlineKeyboardButton(
            f"{prem} {busy} {c['name']} ({c['total_deliveries']} ta)",
            callback_data=f"admin_courier_detail_{c['tg_id']}"
        )])

    buttons.append([InlineKeyboardButton("➕ Kuryer qo'shish", callback_data="add_courier")])
    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")])
    await q.edit_message_text(
        "🚴 <b>Kuryerlar:</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def admin_courier_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    courier_tg_id = int(q.data.split("_")[3])

    conn = get_db()
    c = conn.execute("SELECT * FROM couriers WHERE tg_id=?", (courier_tg_id,)).fetchone()
    conn.close()

    if not c:
        await q.answer("Topilmadi", show_alert=True)
        return

    busy = "🔴 Band" if c["is_busy"] else "🟢 Bo'sh"
    prem = "⚡ Premium" if c["is_premium"] else "Oddiy"
    text = (
        f"🚴 <b>{c['name']}</b>\n\n"
        f"📊 Holat: {busy}\n"
        f"🏷 Tur: {prem}\n"
        f"📦 Yetkazishlar: {c['total_deliveries']}\n"
        f"🆔 Telegram ID: {c['tg_id']}\n"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔥 Ishdan bo'shatish", callback_data=f"admin_courier_fire_{courier_tg_id}")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_couriers")],
    ])
    await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

async def admin_courier_fire(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    courier_tg_id = int(q.data.split("_")[3])

    conn = get_db()
    c = conn.execute("SELECT name FROM couriers WHERE tg_id=?", (courier_tg_id,)).fetchone()
    conn.execute("DELETE FROM couriers WHERE tg_id=?", (courier_tg_id,))
    conn.execute("UPDATE users SET role='customer' WHERE tg_id=?", (courier_tg_id,))
    conn.commit()
    conn.close()

    name = c["name"] if c else str(courier_tg_id)
    await q.edit_message_text(
        f"✅ {name} ishdan bo'shatildi.",
        reply_markup=back_kb("admin_couriers")
    )
    try:
        await context.bot.send_message(
            courier_tg_id,
            "❌ Siz OsonSavdo kuryer tizimidan chiqarildingiz."
        )
    except:
        pass

async def add_courier_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    await q.edit_message_text(
        "🚴 Kuryer ma'lumotlarini kiriting:\n\nFormat: TelegramID|Ism|premium(1/0)\nMasalan: 123456|Ali|0"
    )
    return WAITING_COURIER_NAME

async def got_courier_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        parts = update.message.text.strip().split("|")
        tg_id = int(parts[0])
        name = parts[1]
        is_premium = int(parts[2]) if len(parts) > 2 else 0
    except:
        await update.message.reply_text("❌ Format: TelegramID|Ism|premium(1/0)")
        return WAITING_COURIER_NAME

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO couriers (tg_id, name, is_premium) VALUES (?, ?, ?)",
            (tg_id, name, is_premium)
        )
        conn.execute("INSERT OR IGNORE INTO users (tg_id, full_name, role) VALUES (?, ?, 'courier')", (tg_id, name))
        conn.execute("UPDATE users SET role='courier' WHERE tg_id=?", (tg_id,))
        conn.commit()
        await update.message.reply_text(f"✅ Kuryer {name} qo'shildi!")
        try:
            await context.bot.send_message(tg_id, f"🚴 Siz OsonSavdo kuryeri sifatida qo'shildingiz!\n/start bosing.")
        except:
            pass
    except Exception as e:
        await update.message.reply_text(f"❌ Xatolik: {e}")
    finally:
        conn.close()

    return ConversationHandler.END

# ─── COURIER STATS ─────────────────────────────────────────────────────────────
async def courier_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    orders = conn.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id "
        "WHERE o.courier_tg_id=? ORDER BY o.id DESC LIMIT 10",
        (tg_id,)
    ).fetchall()
    conn.close()

    if not orders:
        await q.edit_message_text("📦 Buyurtmalar yo'q.", reply_markup=back_kb())
        return

    buttons = []
    for o in orders:
        emoji = order_status_emoji(o["status"])
        buttons.append([InlineKeyboardButton(
            f"{emoji} #{1000 + o['id']} — {o['shop_name']}",
            callback_data=f"courier_order_{o['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")])
    await q.edit_message_text(
        "🚴 <b>Mening yetkazishlarim:</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

async def courier_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    conn = get_db()
    courier = conn.execute("SELECT * FROM couriers WHERE tg_id=?", (tg_id,)).fetchone()
    conn.close()

    if not courier:
        await q.edit_message_text("Kuryer topilmadi.", reply_markup=back_kb())
        return

    text = (
        f"📊 <b>Kuryer statistikasi</b>\n\n"
        f"👤 Ism: {courier['name']}\n"
        f"📦 Jami yetkazishlar: {courier['total_deliveries']}\n"
        f"⭐ Reyting: {courier['rating']:.1f}\n"
        f"{'⚡ Premium kuryer' if courier['is_premium'] else '🚴 Oddiy kuryer'}\n"
        f"{'🔴 Band' if courier['is_busy'] else '🟢 Bo\'sh'}\n"
    )

    await q.edit_message_text(text, reply_markup=back_kb(), parse_mode="HTML")

# ─── PROMO CODES (ADMIN) ───────────────────────────────────────────────────────
async def admin_promos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    promos = conn.execute("SELECT * FROM promo_codes ORDER BY id DESC").fetchall()
    conn.close()

    text = "🎫 <b>Promo kodlar:</b>\n\n"
    for p in promos:
        status = "✅" if p["is_active"] else "❌"
        disc = f"{p['discount_value']}%" if p["discount_type"] == "percent" else format_price(p["discount_value"])
        text += f"{status} <code>{p['code']}</code> — {disc} | {p['used_count']}/{p['max_uses']}\n"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Yangi promo", callback_data="create_promo")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")],
    ])
    await q.edit_message_text(text or "Promo kodlar yo'q.", reply_markup=kb, parse_mode="HTML")

async def create_promo_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    context.user_data.pop("new_promo", None)
    await q.edit_message_text(
        "🎫 <b>Yangi promo kod yaratish</b>\n\n"
        "<b>1-qadam</b> / 6\n\n"
        "✏️ Promo kod nomini kiriting:\n"
        "<i>Faqat lotin harflar va raqamlar. Masalan: YOZI25, CHEGIRMA10</i>",
        parse_mode="HTML",
        reply_markup=back_kb("admin_promos")
    )
    return WAITING_PROMO_CODE

async def got_promo_code(update: Update, context: ContextTypes.DEFAULT_TYPE):
    code = update.message.text.strip().upper()
    if len(code) < 2 or len(code) > 20:
        await update.message.reply_text(
            "❌ Kod 2-20 belgi orasida bo'lishi kerak. Qayta kiriting:"
        )
        return WAITING_PROMO_CODE

    context.user_data["new_promo"] = {"code": code}

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 Foiz chegirma (%)", callback_data="promo_type_percent")],
        [InlineKeyboardButton("💵 Aniq summa (so'm)", callback_data="promo_type_fixed")],
    ])
    await update.message.reply_text(
        f"✅ Kod: <b>{code}</b>\n\n"
        f"<b>2-qadam</b> / 6\n\n"
        f"💰 Chegirma turini tanlang:\n\n"
        f"📊 <b>Foiz</b> — buyurtma summasidan % oladi\n"
        f"💵 <b>Aniq summa</b> — har doim belgilangan so'm chegiradi",
        parse_mode="HTML",
        reply_markup=kb
    )
    return WAITING_PROMO_TYPE

async def got_promo_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    dtype = "percent" if q.data == "promo_type_percent" else "fixed"
    context.user_data["new_promo"]["type"] = dtype

    if dtype == "percent":
        label = "📊 Foiz chegirma (%)"
        hint = "Masalan: <code>25</code>  →  25% chegirma"
        unit = "%"
    else:
        label = "💵 Aniq summa (so'm)"
        hint = "Masalan: <code>10000</code>  →  10 000 so'm chegirma"
        unit = "so'm"

    await q.edit_message_text(
        f"✅ Tur: <b>{label}</b>\n\n"
        f"<b>3-qadam</b> / 6\n\n"
        f"🔢 Chegirma miqdorini kiriting ({unit}):\n"
        f"<i>{hint}</i>",
        parse_mode="HTML",
        reply_markup=back_kb("admin_promos")
    )
    return WAITING_PROMO_VALUE

async def got_promo_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        value = float(update.message.text.strip().replace(" ", "").replace(",", ""))
        if value <= 0:
            raise ValueError
        promo = context.user_data.get("new_promo", {})
        if promo.get("type") == "percent" and value > 100:
            await update.message.reply_text("❌ Foiz 100% dan oshmasin. Qayta kiriting:")
            return WAITING_PROMO_VALUE
    except:
        await update.message.reply_text("❌ Noto'g'ri raqam. Qayta kiriting:")
        return WAITING_PROMO_VALUE

    context.user_data["new_promo"]["value"] = value
    dtype = context.user_data["new_promo"]["type"]
    unit = "%" if dtype == "percent" else " so'm"

    await update.message.reply_text(
        f"✅ Chegirma: <b>{value:.0f}{unit}</b>\n\n"
        f"<b>4-qadam</b> / 6\n\n"
        f"💳 Minimal buyurtma summasi (so'm):\n"
        f"<i>Koddan foydalanish uchun buyurtma kamida shu summadan oshishi kerak.\n"
        f"Limit yo'q bo'lsa <code>0</code> kiriting.</i>",
        parse_mode="HTML"
    )
    return WAITING_PROMO_MIN_AMOUNT

async def got_promo_min_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        amount = float(update.message.text.strip().replace(" ", "").replace(",", ""))
        if amount < 0:
            raise ValueError
    except:
        await update.message.reply_text("❌ Noto'g'ri raqam. 0 yoki musbat son kiriting:")
        return WAITING_PROMO_MIN_AMOUNT

    context.user_data["new_promo"]["min_amount"] = amount

    min_text = f"{amount:,.0f} so'mdan yuqori buyurtmalarda" if amount > 0 else "barcha buyurtmalarda"

    await update.message.reply_text(
        f"✅ Minimal summa: <b>{format_price(amount) if amount > 0 else 'Cheklovsiz'}</b>\n\n"
        f"<b>5-qadam</b> / 6\n\n"
        f"👥 Necha marta ishlatish mumkin? (foydalanish limiti)\n"
        f"<i>Masalan: <code>100</code>  →  100 marta ishlatiladi</i>",
        parse_mode="HTML"
    )
    return WAITING_PROMO_LIMIT

async def got_promo_limit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        limit = int(update.message.text.strip())
        if limit <= 0:
            raise ValueError
    except:
        await update.message.reply_text("❌ Noto'g'ri raqam. Butun musbat son kiriting:")
        return WAITING_PROMO_LIMIT

    context.user_data["new_promo"]["limit"] = limit

    await update.message.reply_text(
        f"✅ Limit: <b>{limit} marta</b>\n\n"
        f"<b>6-qadam</b> / 6\n\n"
        f"📅 Necha kun amal qiladi?\n"
        f"<i>Masalan: <code>30</code>  →  30 kun</i>",
        parse_mode="HTML"
    )
    return WAITING_PROMO_DAYS

async def got_promo_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        days = int(update.message.text.strip())
        if days <= 0:
            raise ValueError
    except:
        await update.message.reply_text("❌ Noto'g'ri raqam. Kunlar sonini kiriting:")
        return WAITING_PROMO_DAYS

    promo = context.user_data.pop("new_promo", {})
    code = promo.get("code", "")
    dtype = promo.get("type", "percent")
    value = promo.get("value", 0)
    limit = promo.get("limit", 100)
    min_amount = promo.get("min_amount", 0)
    expires = (datetime.now() + timedelta(days=days)).isoformat()
    unit = "%" if dtype == "percent" else " so'm"

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO promo_codes (code, discount_type, discount_value, max_uses, expires_at, min_order_amount) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (code, dtype, value, limit, expires, min_amount)
        )
        conn.commit()
        min_text = format_price(min_amount) if min_amount > 0 else "Cheklovsiz"
        await update.message.reply_text(
            f"🎉 <b>Promo kod yaratildi!</b>\n\n"
            f"🎫 Kod: <code>{code}</code>\n"
            f"💰 Chegirma: <b>{value:.0f}{unit}</b>\n"
            f"💳 Minimal summa: <b>{min_text}</b>\n"
            f"👥 Limit: <b>{limit} marta</b>\n"
            f"📅 Muddat: <b>{days} kun</b> ({expires[:10]} gacha)",
            parse_mode="HTML",
            reply_markup=back_kb("admin_promos")
        )
    except:
        await update.message.reply_text(
            f"❌ <b>{code}</b> kodi allaqachon mavjud!",
            parse_mode="HTML",
            reply_markup=back_kb("admin_promos")
        )
    finally:
        conn.close()

    return ConversationHandler.END

# ─── ADMIN COMMISSION ──────────────────────────────────────────────────────────
async def admin_commission(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    current = get_setting("commission_percent", "10")
    await q.edit_message_text(
        f"💰 Joriy komissiya: <b>{current}%</b>\n\nYangi foizni kiriting (0-50):",
        reply_markup=back_kb("admin_panel"),
        parse_mode="HTML"
    )
    return WAITING_COMMISSION

async def got_commission(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        pct = float(update.message.text.strip())
        if not 0 <= pct <= 50:
            raise ValueError
    except:
        await update.message.reply_text("❌ 0 dan 50 gacha raqam kiriting.")
        return WAITING_COMMISSION

    set_setting("commission_percent", str(pct))
    await update.message.reply_text(
        f"✅ Komissiya {pct}% ga o'zgartirildi!",
        reply_markup=back_kb("admin_panel")
    )
    return ConversationHandler.END

# ─── ADMIN TICKETS ─────────────────────────────────────────────────────────────
async def admin_tickets(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    conn = get_db()
    tickets = conn.execute("SELECT * FROM tickets ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()

    if not tickets:
        await q.edit_message_text("🎟 Ticketlar yo'q.", reply_markup=back_kb("admin_panel"))
        return

    buttons = []
    for t in tickets:
        status = "🟢" if t["status"] == "open" else "🔴"
        buttons.append([InlineKeyboardButton(
            f"{status} #SUP-{1000 + t['id']} — {t['subject'][:30]}",
            callback_data=f"ticket_reply_{t['id']}"
        )])

    buttons.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")])
    await q.edit_message_text(
        "🎟 <b>Ticketlar:</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode="HTML"
    )

# ─── ADMIN EXCEL REPORT ────────────────────────────────────────────────────────
async def admin_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id

    if not EXCEL_AVAILABLE:
        await q.edit_message_text("❌ openpyxl o'rnatilmagan.", reply_markup=back_kb("admin_panel"))
        return

    conn = get_db()
    orders = conn.execute(
        "SELECT o.*, s.name as shop_name FROM orders o LEFT JOIN shops s ON o.shop_id=s.id ORDER BY o.id DESC"
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"

    headers = ["Buyurtma ID", "Do'kon", "Manzil", "Jami", "Komissiya", "Holat", "Sana"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    for row, o in enumerate(orders, 2):
        ws.append([
            f"#{1000 + o['id']}",
            o["shop_name"],
            o["address"],
            o["total"],
            o["commission"],
            order_status_text(o["status"]),
            o["created_at"][:10]
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await context.bot.send_document(
        tg_id,
        document=InputFile(buf, filename=f"admin_hisobot_{datetime.now().strftime('%Y%m%d')}.xlsx"),
        caption="📊 Admin hisobot"
    )

# ─── BROADCAST ─────────────────────────────────────────────────────────────────
async def admin_broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    await q.edit_message_text("📢 Xabarni kiriting (barcha foydalanuvchilarga yuboriladi):")
    return WAITING_BROADCAST

async def got_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message.text.strip()
    conn = get_db()
    users = conn.execute("SELECT tg_id FROM users").fetchall()
    conn.close()

    sent = 0
    for u in users:
        try:
            await context.bot.send_message(u["tg_id"], f"📢 <b>OsonSavdo xabari:</b>\n\n{msg}", parse_mode="HTML")
            sent += 1
        except:
            pass

    await update.message.reply_text(f"✅ Xabar {sent}/{len(users)} foydalanuvchiga yuborildi!")
    return ConversationHandler.END

# ─── ADMIN SETTINGS ────────────────────────────────────────────────────────────
async def admin_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    commission = get_setting("commission_percent", "10")
    premium_fee = get_setting("premium_courier_fee", "15000")
    ref_bonus = get_setting("referral_bonus", "5000")
    phones_raw = get_setting("order_phones", "")
    phones = json.loads(phones_raw) if phones_raw else []

    phones_text = ""
    if phones:
        phones_text = "\n\n📞 <b>Telefon raqamlar:</b>\n"
        for i, p in enumerate(phones):
            shop_n = f" ({p.get('shop_name','')})" if p.get('shop_name') else ""
            phones_text += f"{i+1}. {p['name']}{shop_n}: <code>{p['phone']}</code>\n"
    else:
        phones_text = "\n\n📞 Telefon raqamlar: <i>yo'q</i>"

    text = (
        f"⚙️ <b>Platform sozlamalari</b>\n\n"
        f"💰 Komissiya: {commission}%\n"
        f"⚡ Premium kuryer: {format_price(float(premium_fee))}\n"
        f"🔗 Referal bonus: {format_price(float(ref_bonus))}"
        f"{phones_text}"
    )

    phone_buttons = []
    for i, p in enumerate(phones):
        phone_buttons.append([InlineKeyboardButton(
            f"🗑 {p['name']}: {p['phone']}",
            callback_data=f"del_phone_{i}"
        )])

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("💰 Komissiya o'zgartirish", callback_data="admin_commission")],
        [InlineKeyboardButton("➕ Telefon raqam qo'shish", callback_data="add_phone")],
        *phone_buttons,
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="admin_panel")],
    ])
    await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

# ─── OPERATOR PANEL ────────────────────────────────────────────────────────────
async def operator_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 Buyurtma kiritish", callback_data="operator_new_order")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="main_menu")],
    ])
    await q.edit_message_text("🎙 <b>Operator Panel</b>", reply_markup=kb, parse_mode="HTML")

async def operator_new_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    await q.edit_message_text(
        "📝 Buyurtmani matn shaklida kiriting:\n\nFormat: FoydalanuvchiID|Mahsulotlar|Manzil"
    )
    return WAITING_OPERATOR_ORDER

async def got_operator_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    tg_id = update.effective_user.id

    conn = get_db()
    conn.execute(
        "INSERT INTO operator_orders (operator_tg_id, raw_request) VALUES (?, ?)",
        (tg_id, text)
    )
    conn.commit()
    conn.close()

    try:
        await context.bot.send_message(
            ADMIN_ID,
            f"🎙 <b>Operator buyurtmasi</b>\n\n{text}\n\nOperator: {update.effective_user.full_name}",
            parse_mode="HTML"
        )
    except:
        pass

    await update.message.reply_text("✅ Buyurtma adminiga yuborildi!", reply_markup=back_kb())
    return ConversationHandler.END

# ─── NOOP ──────────────────────────────────────────────────────────────────────
async def noop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

# ─── MAIN MENU CALLBACK ────────────────────────────────────────────────────────
async def main_menu_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    tg_id = update.effective_user.id
    role = get_user_role(tg_id)
    platform = get_setting("platform_name", "OsonSavdo")

    conn = get_db()
    shops = conn.execute("SELECT * FROM shops WHERE status='approved' ORDER BY rating DESC").fetchall()
    conn.close()

    cart = get_cart(context)
    cart_count = sum(v["qty"] for v in cart.values())
    cart_label = f"🛒 Savat ({cart_count})" if cart_count else "🛒 Savat"

    if role in ("customer", "admin", "operator"):
        if not shops:
            text = f"🛍 <b>{platform}</b>\n\n🏪 Hozircha do'konlar yo'q."
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(cart_label, callback_data="cart_view")],
                [InlineKeyboardButton("☰ Ko'proq", callback_data="more_menu")],
            ])
        else:
            text = f"🛍 <b>{platform}</b> — Do'konlar\n\n"
            buttons = []
            for s in shops:
                is_open = s["is_open"] if "is_open" in s.keys() else 1
                status_icon = "🟢" if is_open else "🔴"
                rating = f"⭐{s['rating']:.1f}" if s["rating"] else "⭐"
                buttons.append([InlineKeyboardButton(
                    f"{status_icon} {s['name']} {rating} | 🚚{format_price(s['delivery_price'])}",
                    callback_data=f"shop_{s['id']}"
                )])
            buttons.append([
                InlineKeyboardButton(cart_label, callback_data="cart_view"),
                InlineKeyboardButton("☰ Ko'proq", callback_data="more_menu"),
            ])
            kb = InlineKeyboardMarkup(buttons)
    else:
        role_names = {"shop_owner": "Do'kon egasi", "courier": "Kuryer", "admin": "Admin", "operator": "Operator"}
        text = (
            f"👋 <b>{q.from_user.full_name}</b>!\n"
            f"🛍 <b>{platform}</b>\n\n"
            f"👤 Rolingiz: <b>{role_names.get(role, role)}</b>"
        )
        kb = main_menu_kb(role)

    await q.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

# ─── APPLICATION ───────────────────────────────────────────────────────────────
def main():
    init_db()
    migrate_db()

    async def on_startup(application):
        async def subscription_loop():
            while True:
                await asyncio.sleep(86400)
                try:
                    class FakeContext:
                        bot = application.bot
                    await check_subscriptions(FakeContext())
                except Exception as e:
                    logger.error(f"Obuna tekshiruvida xato: {e}")

        async def shop_hours_loop():
            # Darhol bir marta tekshir
            try:
                await auto_open_close_shops(application.bot)
            except Exception as e:
                logger.error(f"Ish vaqti tekshiruvida xato: {e}")
            while True:
                await asyncio.sleep(60)  # Har 1 daqiqada
                try:
                    await auto_open_close_shops(application.bot)
                except Exception as e:
                    logger.error(f"Ish vaqti tekshiruvida xato: {e}")

        asyncio.create_task(subscription_loop())
        asyncio.create_task(shop_hours_loop())

    persistence = PicklePersistence(filepath="/data/oson_savdo_persistence")
    app = Application.builder().token(BOT_TOKEN).persistence(persistence).post_init(on_startup).build()

    # Conversation handlers
    checkout_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(checkout, pattern="^checkout$")],
        states={WAITING_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_address)]},
        fallbacks=[],
        per_message=False,
    )

    payment_screenshot_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(courier_type_select, pattern="^courier_(standard|premium)$")],
        states={WAITING_PAYMENT_SCREENSHOT: [MessageHandler(filters.PHOTO | filters.Document.ALL, got_payment_screenshot)]},
        fallbacks=[],
        per_message=False,
    )

    promo_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(promo_enter, pattern="^promo_enter$")],
        states={WAITING_PROMO: [MessageHandler(filters.TEXT & ~filters.COMMAND, promo_apply)]},
        fallbacks=[],
        per_message=False,
    )

    shop_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_shop_start, pattern="^add_shop$")],
        states={
            WAITING_SHOP_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_shop_name)],
            WAITING_SHOP_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_shop_desc)],
        },
        fallbacks=[],
        per_message=False,
    )

    product_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_product_start, pattern="^add_product$")],
        states={
            WAITING_PRODUCT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_product_name)],
            WAITING_PRODUCT_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_product_desc)],
            WAITING_PRODUCT_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_product_price)],
            WAITING_PRODUCT_PHOTO: [
                MessageHandler(filters.PHOTO, got_product_photo),
                MessageHandler(filters.TEXT & filters.Regex(r'^/skip$'), got_product_photo),
            ],
        },
        fallbacks=[],
        per_message=False,
    )

    review_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(give_rating, pattern=r"^give_rating_\d+_\d+$")],
        states={WAITING_REVIEW: [MessageHandler(filters.TEXT, got_review)]},
        fallbacks=[],
        per_message=False,
    )

    ticket_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(ticket_open, pattern="^ticket_open$")],
        states={WAITING_TICKET_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_ticket_msg)]},
        fallbacks=[],
        per_message=False,
    )

    ticket_reply_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(ticket_reply_prompt, pattern=r"^ticket_reply_\d+$")],
        states={WAITING_TICKET_REPLY: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_ticket_reply)]},
        fallbacks=[],
        per_message=False,
    )

    courier_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_courier_start, pattern="^add_courier$")],
        states={WAITING_COURIER_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_courier_info)]},
        fallbacks=[],
        per_message=False,
    )

    promo_create_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(create_promo_start, pattern="^create_promo$")],
        states={
            WAITING_PROMO_CODE:       [MessageHandler(filters.TEXT & ~filters.COMMAND, got_promo_code)],
            WAITING_PROMO_TYPE:       [CallbackQueryHandler(got_promo_type, pattern="^promo_type_")],
            WAITING_PROMO_VALUE:      [MessageHandler(filters.TEXT & ~filters.COMMAND, got_promo_value)],
            WAITING_PROMO_MIN_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_promo_min_amount)],
            WAITING_PROMO_LIMIT:      [MessageHandler(filters.TEXT & ~filters.COMMAND, got_promo_limit)],
            WAITING_PROMO_DAYS:       [MessageHandler(filters.TEXT & ~filters.COMMAND, got_promo_days)],
        },
        fallbacks=[],
        per_message=False,
    )

    commission_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(admin_commission, pattern="^admin_commission$")],
        states={WAITING_COMMISSION: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_commission)]},
        fallbacks=[],
        per_message=False,
    )

    broadcast_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(admin_broadcast_start, pattern="^admin_broadcast$")],
        states={WAITING_BROADCAST: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_broadcast)]},
        fallbacks=[],
        per_message=False,
    )

    delivery_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(set_delivery_price, pattern=r"^set_delivery_\d+$")],
        states={WAITING_DELIVERY_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_delivery_price)]},
        fallbacks=[],
        per_message=False,
    )

    hours_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(set_work_hours, pattern=r"^set_hours_\d+$")],
        states={WAITING_WORK_HOURS: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_work_hours)]},
        fallbacks=[],
        per_message=False,
    )

    edit_price_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(edit_product_price, pattern=r"^edit_price_\d+$")],
        states={WAITING_PRODUCT_EDIT_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_edit_price)]},
        fallbacks=[],
        per_message=False,
    )

    discount_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(set_product_discount, pattern=r"^set_discount_\d+$")],
        states={WAITING_PRODUCT_DISCOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_product_discount)]},
        fallbacks=[],
        per_message=False,
    )

    operator_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(operator_new_order, pattern="^operator_new_order$")],
        states={WAITING_OPERATOR_ORDER: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_operator_order)]},
        fallbacks=[],
        per_message=False,
    )

    job_shop_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(job_shop_owner_start, pattern="^job_shop_owner$")],
        states={
            WAITING_JOB_SHOP_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_job_shop_name)],
            WAITING_JOB_SHOP_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_job_shop_desc)],
        },
        fallbacks=[],
        per_message=False,
    )

    admin_add_shop_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(admin_add_shop_start, pattern="^admin_add_shop$")],
        states={
            WAITING_ADMIN_SHOP_OWNER: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_admin_shop_owner)],
            WAITING_ADMIN_SHOP_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_admin_shop_name)],
            WAITING_ADMIN_SHOP_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_admin_shop_desc)],
        },
        fallbacks=[],
        per_message=False,
    )

    # Do'kon egasi to'lov tizimi
    payment_settings_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(set_payment_settings, pattern=r"^set_payment_\d+$")],
        states={
            WAITING_CARD_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_card_number)],
            WAITING_CARD_HOLDER: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_card_holder)],
        },
        fallbacks=[],
        per_message=False,
    )

    # Admin to'lov tizimi
    admin_payment_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(admin_set_payment_start, pattern=r"^admin_set_payment_\d+$")],
        states={
            WAITING_CARD_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_got_card_number)],
            WAITING_CARD_HOLDER: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_got_card_holder)],
        },
        fallbacks=[],
        per_message=False,
    )

    # Add handlers
    for conv in [
        admin_add_shop_conv,
        checkout_conv, payment_screenshot_conv, promo_conv, shop_conv, product_conv,
        review_conv, ticket_conv, ticket_reply_conv, courier_conv, promo_create_conv,
        commission_conv, broadcast_conv, delivery_conv, hours_conv, edit_price_conv,
        discount_conv, operator_conv, job_shop_conv,
        payment_settings_conv, admin_payment_conv,
    ]:
        app.add_handler(conv)

    sub_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(admin_sub_settings, pattern=r"^admin_sub_\d+$")],
        states={WAITING_SUB_PERCENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_sub_percent)]},
        fallbacks=[],
        per_message=False,
    )
    app.add_handler(sub_conv)

    # Admin telefon raqam qo'shish
    admin_phone_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_phone_start, pattern="^add_phone$")],
        states={
            WAITING_PHONE_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_admin_phone_name)],
            WAITING_PHONE_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_admin_phone_number)],
        },
        fallbacks=[],
        per_message=False,
    )
    app.add_handler(admin_phone_conv, group=2)

    # Do'kon egasi telefon raqam qo'shish
    shop_phone_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(shop_phone_start, pattern=r"^shop_phone_\d+$")],
        states={
            WAITING_PHONE_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_shop_phone_name)],
            WAITING_PHONE_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_shop_phone_number)],
        },
        fallbacks=[],
        per_message=False,
    )
    app.add_handler(shop_phone_conv, group=3)

    # Do'kon egasi telefon orqali buyurtma kiritish
    tel_order_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(tel_order_new, pattern=r"^tel_order_new_\d+$")],
        states={
            WAITING_TEL_ORDER_ITEMS: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_tel_order_items)],
            WAITING_TEL_ORDER_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, got_tel_order_address)],
        },
        fallbacks=[],
        per_message=False,
    )
    app.add_handler(tel_order_conv, group=1)

    app.add_handler(CommandHandler("start", start))

    # Callback handlers
    app.add_handler(CallbackQueryHandler(main_menu_cb, pattern="^main_menu$"))
    app.add_handler(CallbackQueryHandler(more_menu, pattern="^more_menu$"))
    app.add_handler(CallbackQueryHandler(shops_list, pattern="^shops_list$"))
    app.add_handler(CallbackQueryHandler(shop_detail, pattern=r"^shop_(?!call_|settings_|orders|report|confirm_|reject_|rating_)\d+$"))
    app.add_handler(CallbackQueryHandler(products_list, pattern=r"^products_\d+$"))
    app.add_handler(CallbackQueryHandler(product_detail, pattern=r"^product_\d+$"))
    app.add_handler(CallbackQueryHandler(add_to_cart, pattern=r"^add_cart_\d+$"))
    app.add_handler(CallbackQueryHandler(cart_view, pattern="^cart_view$"))
    app.add_handler(CallbackQueryHandler(cart_update, pattern=r"^cart_(inc|dec|del)_\d+$"))
    app.add_handler(CallbackQueryHandler(cart_clear, pattern="^cart_clear$"))
    app.add_handler(CallbackQueryHandler(payment_method, pattern="^pay_(card|cash|bonus)$"))
    app.add_handler(CallbackQueryHandler(my_orders, pattern="^my_orders$"))
    app.add_handler(CallbackQueryHandler(order_detail, pattern=r"^order_detail_\d+$"))
    app.add_handler(CallbackQueryHandler(reorder, pattern=r"^reorder_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_confirm_order, pattern=r"^admin_confirm_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_reject_order, pattern=r"^admin_reject_\d+$"))
    app.add_handler(CallbackQueryHandler(shop_confirm_order, pattern=r"^shop_confirm_\d+$"))
    app.add_handler(CallbackQueryHandler(shop_reject_order, pattern=r"^shop_reject_\d+$"))
    app.add_handler(CallbackQueryHandler(courier_accept, pattern=r"^courier_accept_\d+$"))
    app.add_handler(CallbackQueryHandler(courier_reject, pattern=r"^courier_reject_\d+$"))
    app.add_handler(CallbackQueryHandler(courier_delivered, pattern=r"^courier_delivered_\d+$"))
    app.add_handler(CallbackQueryHandler(profile_view, pattern="^profile$"))
    app.add_handler(CallbackQueryHandler(favorites_view, pattern="^favorites$"))
    app.add_handler(CallbackQueryHandler(add_favorite, pattern=r"^fav_\d+$"))
    app.add_handler(CallbackQueryHandler(remove_favorite, pattern=r"^unfav_\d+$"))
    app.add_handler(CallbackQueryHandler(rate_order, pattern=r"^rate_order_\d+$"))
    app.add_handler(CallbackQueryHandler(give_rating, pattern=r"^give_rating_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(rate_shop, pattern=r"^rate_shop_\d+$"))
    app.add_handler(CallbackQueryHandler(give_shop_rating, pattern=r"^shop_rating_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(referral_view, pattern="^referral$"))
    app.add_handler(CallbackQueryHandler(my_shop, pattern="^my_shop$"))
    app.add_handler(CallbackQueryHandler(shop_settings, pattern=r"^shop_settings_\d+$"))
    app.add_handler(CallbackQueryHandler(set_payment_settings, pattern=r"^set_payment_\d+$"))
    app.add_handler(CallbackQueryHandler(owner_products, pattern=r"^owner_products_\d+$"))
    app.add_handler(CallbackQueryHandler(owner_product_detail, pattern=r"^owner_prod_\d+$"))
    app.add_handler(CallbackQueryHandler(toggle_product, pattern=r"^toggle_prod_\d+$"))
    app.add_handler(CallbackQueryHandler(shop_orders, pattern="^shop_orders$"))
    app.add_handler(CallbackQueryHandler(shop_report, pattern="^shop_report$"))
    app.add_handler(CallbackQueryHandler(admin_panel, pattern="^admin_panel$"))
    app.add_handler(CallbackQueryHandler(admin_shops, pattern="^admin_shops$"))
    app.add_handler(CallbackQueryHandler(admin_shop_detail, pattern=r"^admin_shop_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_shop_approve, pattern=r"^admin_shop_approve_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_shop_reject, pattern=r"^admin_shop_reject_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_users, pattern="^admin_users$"))
    app.add_handler(CallbackQueryHandler(admin_orders, pattern="^admin_orders$"))
    app.add_handler(CallbackQueryHandler(admin_order_detail, pattern=r"^admin_order_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_couriers, pattern="^admin_couriers$"))
    app.add_handler(CallbackQueryHandler(admin_promos, pattern="^admin_promos$"))
    app.add_handler(CallbackQueryHandler(admin_tickets, pattern="^admin_tickets$"))
    app.add_handler(CallbackQueryHandler(admin_excel, pattern="^admin_excel$"))
    app.add_handler(CallbackQueryHandler(admin_settings, pattern="^admin_settings$"))
    app.add_handler(CallbackQueryHandler(operator_panel, pattern="^operator_panel$"))
    app.add_handler(CallbackQueryHandler(courier_orders, pattern="^courier_orders$"))
    app.add_handler(CallbackQueryHandler(courier_stats, pattern="^courier_stats$"))
    app.add_handler(CallbackQueryHandler(noop, pattern="^noop$"))
    app.add_handler(CallbackQueryHandler(job_apply, pattern="^job_apply$"))
    app.add_handler(CallbackQueryHandler(job_courier_apply, pattern="^job_courier$"))
    app.add_handler(CallbackQueryHandler(admin_job_requests, pattern="^admin_job_requests$"))
    app.add_handler(CallbackQueryHandler(admin_courier_approve, pattern=r"^admin_courier_approve_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_courier_reject, pattern=r"^admin_courier_reject_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_courier_detail, pattern=r"^admin_courier_detail_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_courier_fire, pattern=r"^admin_courier_fire_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_shop_delete, pattern=r"^admin_shop_delete_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_set_payment_start, pattern=r"^admin_set_payment_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_sub_settings, pattern=r"^admin_sub_\d+$"))
    app.add_handler(CallbackQueryHandler(toggle_shop, pattern=r"^toggle_shop_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_toggle_shop, pattern=r"^admin_toggle_shop_\d+$"))
    app.add_handler(CallbackQueryHandler(del_phone_cb, pattern=r"^del_phone_\d+$"))
    app.add_handler(CallbackQueryHandler(shop_del_phone_cb, pattern=r"^shop_del_phone_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(tel_order_confirm, pattern=r"^tel_confirm_\d+$"))
    app.add_handler(CallbackQueryHandler(phone_order_view, pattern="^phone_order$"))
    app.add_handler(CallbackQueryHandler(shop_call_view, pattern=r"^shop_call_\d+$"))

    logger.info("🚀 OsonSavdo Bot ishga tushdi!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
